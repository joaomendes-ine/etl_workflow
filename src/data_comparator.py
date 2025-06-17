"""
Módulo de Comparação Inteligente de Dados Excel - Versão Avançada
Compara ficheiros Excel com estruturas de tabela cruzada (crosstab/pivot)
Suporta formatação visual, células mescladas, valores apresentados e relatórios destacados.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.formatting.rule import FormulaRule
import re
from typing import Dict, List, Tuple, Any, Optional, Set
from datetime import datetime
import logging
from difflib import SequenceMatcher
from copy import copy
from src.utils import ensure_directory_exists
from openpyxl.comments import Comment
from colorama import Fore, Style, init
import openpyxl


class DataComparator:
    """
    Classe principal para comparação inteligente de ficheiros Excel.
    Especializada em estruturas de tabela cruzada com normalização avançada.
    """
    
    def __init__(self, logger: logging.Logger):
        """
        Inicializa o comparador de dados.
        
        Args:
            logger: Logger configurado para registo de operações
        """
        self.logger = logger
        self.comparison_results = []
        self.summary_stats = {}
        self.dimension_mapping = {}
        
        # Configurações de tolerância para comparação numérica
        self.numeric_tolerance = 1.0  # Tolerância de ±1 para números publicados (arredondados)
        self.percentage_tolerance = 0.001
        
        # Padrões para normalização de dimensões
        self.dimension_patterns = {
            r'\(em\s+branco\)': 'Total',
            r'de\s+(\d+)\s+a\s+(\d+)': r'\1 - \2',
            r'(\d+)\s*anos?\s*ou\s*mais': r'\1+',
            r'menos\s+de\s+(\d+)\s*anos?': r'< \1',
            r'^\s*-\s*$': 'Não especificado',
            r'n\.?\s*d\.?\s*': 'Não disponível'
        }
        
        # NOVA: Mapeamento de equivalências semânticas descobertas
        self.semantic_equivalents = {
            "****": "4", "***": "3", "**": "2", "*": "1",
            "Total": "(Em branco)", "(Em branco)": "Total", 
            "(em branco)": "Total", "total": "Total",
            "Estabelecimentos hoteleiros": "Hotelaria",
            "Hotelaria": "Estabelecimentos hoteleiros",
            "Hotéis": "Hotel", "Hotel": "Hotéis",
            "Pensões": "Pensão", "Pensão": "Pensões",
            "Pousadas": "Pousada", "Pousada": "Pousadas",
            "Motéis": "Motel", "Motel": "Motéis",
            "Hotéis-apartamentos": "Hotel-apartamento", 
            "Hotel-apartamento": "Hotéis-apartamentos",
            "n.º": "número", "número": "n.º"
        }
    
    def get_available_files(self) -> Tuple[List[str], Dict[str, List[str]]]:
        """
        Obtém lista de ficheiros disponíveis para comparação.
        
        Returns:
            Tupla com (ficheiros publicados, ficheiros recriados por pasta)
        """
        published_files = []
        recreated_files = {}
        
        # Procura ficheiros publicados em dataset/comparison/
        comparison_dir = "dataset/comparison"
        if os.path.exists(comparison_dir):
            for file in os.listdir(comparison_dir):
                if file.lower().endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    published_files.append(file)
        
        # Procura ficheiros recriados em result/validation/
        validation_dir = "result/validation"
        if os.path.exists(validation_dir):
            for folder in os.listdir(validation_dir):
                folder_path = os.path.join(validation_dir, folder)
                if os.path.isdir(folder_path):
                    folder_files = []
                    
                    # Verifica subpastas series e quadros
                    for subfolder in ['series', 'quadros']:
                        subfolder_path = os.path.join(folder_path, subfolder)
                        if os.path.exists(subfolder_path):
                            for file in os.listdir(subfolder_path):
                                if file.lower().endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                                    folder_files.append(os.path.join(subfolder, file))
                    
                    # Verifica ficheiros na pasta principal
                    for file in os.listdir(folder_path):
                        file_path = os.path.join(folder_path, file)
                        if os.path.isfile(file_path) and file.lower().endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                            folder_files.append(file)
                    
                    if folder_files:
                        recreated_files[folder] = folder_files
        
        return published_files, recreated_files
    
    def select_files_interactively(self) -> Tuple[Optional[str], Optional[str]]:
        """
        Interface interativa para seleção de ficheiros.
        
        Returns:
            Tupla com (caminho_ficheiro_publicado, caminho_ficheiro_recriado)
        """
        from colorama import Fore, Style
        
        published_files, recreated_files = self.get_available_files()
        
        if not published_files:
            print(f"{Fore.RED}Nenhum ficheiro encontrado em dataset/comparison/{Style.RESET_ALL}")
            return None, None
        
        if not recreated_files:
            print(f"{Fore.RED}Nenhum ficheiro encontrado em result/validation/{Style.RESET_ALL}")
            return None, None
        
        # Seleção do ficheiro publicado
        print(f"\n{Fore.GREEN}[Seleção de Ficheiro Publicado]{Style.RESET_ALL}")
        print(f"Ficheiros disponíveis em dataset/comparison/:")
        for i, file in enumerate(published_files, 1):
            print(f"  {Fore.WHITE}{i}.{Style.RESET_ALL} {file}")
        
        while True:
            try:
                choice = input(f"\n{Fore.GREEN}>>{Style.RESET_ALL} Escolha o ficheiro publicado (1-{len(published_files)}): ")
                if choice == '0':
                    return None, None
                
                pub_idx = int(choice) - 1
                if 0 <= pub_idx < len(published_files):
                    published_file = os.path.join("dataset/comparison", published_files[pub_idx])
                    break
                else:
                    print(f"{Fore.RED}Número inválido. Tente novamente.{Style.RESET_ALL}")
            except ValueError:
                print(f"{Fore.RED}Entrada inválida. Digite um número.{Style.RESET_ALL}")
        
        # Seleção do ficheiro recriado
        print(f"\n{Fore.GREEN}[Seleção de Ficheiro Recriado]{Style.RESET_ALL}")
        print(f"Pastas disponíveis em result/validation/:")
        
        folder_list = sorted(recreated_files.keys(), key=lambda x: int(x) if x.isdigit() else float('inf'))
        for i, folder in enumerate(folder_list, 1):
            file_count = len(recreated_files[folder])
            print(f"  {Fore.WHITE}{i}.{Style.RESET_ALL} Pasta {folder} ({file_count} ficheiro{'s' if file_count != 1 else ''})")
        
        while True:
            try:
                choice = input(f"\n{Fore.GREEN}>>{Style.RESET_ALL} Escolha a pasta (1-{len(folder_list)}): ")
                if choice == '0':
                    return None, None
                
                folder_idx = int(choice) - 1
                if 0 <= folder_idx < len(folder_list):
                    selected_folder = folder_list[folder_idx]
                    break
                else:
                    print(f"{Fore.RED}Número inválido. Tente novamente.{Style.RESET_ALL}")
            except ValueError:
                print(f"{Fore.RED}Entrada inválida. Digite um número.{Style.RESET_ALL}")
        
        # Seleção do ficheiro específico na pasta
        folder_files = recreated_files[selected_folder]
        if len(folder_files) == 1:
            recreated_file = os.path.join("result/validation", selected_folder, folder_files[0])
        else:
            print(f"\n{Fore.GREEN}[Seleção de Ficheiro na Pasta {selected_folder}]{Style.RESET_ALL}")
            for i, file in enumerate(folder_files, 1):
                print(f"  {Fore.WHITE}{i}.{Style.RESET_ALL} {file}")
            
            while True:
                try:
                    choice = input(f"\n{Fore.GREEN}>>{Style.RESET_ALL} Escolha o ficheiro (1-{len(folder_files)}): ")
                    if choice == '0':
                        return None, None
                    
                    file_idx = int(choice) - 1
                    if 0 <= file_idx < len(folder_files):
                        recreated_file = os.path.join("result/validation", selected_folder, folder_files[file_idx])
                        break
                    else:
                        print(f"{Fore.RED}Número inválido. Tente novamente.{Style.RESET_ALL}")
                except ValueError:
                    print(f"{Fore.RED}Entrada inválida. Digite um número.{Style.RESET_ALL}")
        
        self.logger.info(f"Ficheiros selecionados: {published_file} vs {recreated_file}")
        return published_file, recreated_file
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Obtém nomes das folhas de um ficheiro Excel.
        
        Args:
            file_path: Caminho do ficheiro Excel
            
        Returns:
            Lista com nomes das folhas
        """
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            self.logger.error(f"Erro ao ler folhas de {file_path}: {e}")
            return []
    
    def select_sheets_interactively(self, published_file: str, recreated_file: str) -> List[str]:
        """
        Interface para seleção de folhas a comparar.
        
        Args:
            published_file: Caminho do ficheiro publicado
            recreated_file: Caminho do ficheiro recriado
            
        Returns:
            Lista com nomes das folhas a comparar
        """
        from colorama import Fore, Style
        
        pub_sheets = set(self.get_sheet_names(published_file))
        rec_sheets = set(self.get_sheet_names(recreated_file))
        
        common_sheets = sorted(pub_sheets.intersection(rec_sheets))
        
        if not common_sheets:
            print(f"{Fore.RED}Nenhuma folha comum encontrada entre os ficheiros.{Style.RESET_ALL}")
            return []
        
        print(f"\n{Fore.GREEN}[Seleção de Folhas para Comparação]{Style.RESET_ALL}")
        print(f"Folhas comuns encontradas:")
        for i, sheet in enumerate(common_sheets, 1):
            print(f"  {Fore.WHITE}{i}.{Style.RESET_ALL} {sheet}")
        
        print(f"  {Fore.WHITE}T.{Style.RESET_ALL} Todas as folhas")
        print(f"  {Fore.WHITE}0.{Style.RESET_ALL} Cancelar")
        
        while True:
            choice = input(f"\n{Fore.GREEN}>>{Style.RESET_ALL} Escolha as folhas (ex: 1,3,5 ou T para todas): ").strip().lower()
            
            if choice == '0':
                return []
            elif choice == 't':
                return common_sheets
            else:
                try:
                    indices = [int(x.strip()) - 1 for x in choice.split(',') if x.strip()]
                    selected_sheets = [common_sheets[i] for i in indices if 0 <= i < len(common_sheets)]
                    
                    if selected_sheets:
                        return selected_sheets
                    else:
                        print(f"{Fore.RED}Nenhuma folha válida selecionada. Tente novamente.{Style.RESET_ALL}")
                except (ValueError, IndexError):
                    print(f"{Fore.RED}Entrada inválida. Use números separados por vírgula.{Style.RESET_ALL}")
    
    def normalize_value(self, value: Any) -> Optional[float]:
        """
        Normaliza valores numéricos com tratamento robusto para diferentes formatos.
        
        Args:
            value: Valor a normalizar
            
        Returns:
            Valor numérico normalizado ou None se não for número
        """
        if pd.isna(value) or value is None:
            return None
        
        # Se já é número
        if isinstance(value, (int, float)):
            return float(value)
        
        # Converte para string e remove espaços
        str_value = str(value).strip()
        
        # Verifica se é vazio
        if not str_value or str_value in ['-', '']:
            return None
        
        # Preserva o valor original para debug
        original_value = str_value
        
        try:
            # Remove espaços de todos os tipos (incluindo não-quebráveis)
            str_value = re.sub(r'\s+', '', str_value)
            
            # Detecta formato baseado na posição de vírgulas e pontos
            comma_count = str_value.count(',')
            dot_count = str_value.count('.')
            
            if comma_count == 0 and dot_count <= 1:
                # Formato simples: só pontos como decimal
                pass
            elif comma_count == 1 and dot_count == 0:
                # Formato português: vírgula como decimal
                str_value = str_value.replace(',', '.')
            elif comma_count > 0 and dot_count > 0:
                # Formato misto - determinar qual é decimal
                last_comma = str_value.rfind(',')
                last_dot = str_value.rfind('.')
                
                if last_comma > last_dot:
                    # Vírgula é decimal, pontos são separadores de milhares
                    str_value = str_value.replace('.', '').replace(',', '.')
                else:
                    # Ponto é decimal, vírgulas são separadores de milhares
                    str_value = str_value.replace(',', '')
            elif comma_count > 1:
                # Múltiplas vírgulas como separadores de milhares
                str_value = str_value.replace(',', '')
            elif dot_count > 1:
                # Múltiplos pontos como separadores de milhares
                # Se o último segmento tem 1-2 dígitos, trata como decimal
                parts = str_value.split('.')
                if len(parts[-1]) <= 2:
                    str_value = ''.join(parts[:-1]) + '.' + parts[-1]
                else:
                    str_value = str_value.replace('.', '')
            
            # Remove caracteres não numéricos restantes (preserva sinal e decimal)
            str_value = re.sub(r'[^\d.\-+]', '', str_value)
            
            result = float(str_value)
            return result
        except (ValueError, TypeError):
            self.logger.debug(f"Valor não numérico ignorado: '{original_value}' -> '{str_value}'")
            return None
    
    def normalize_dimension_label(self, label: Any) -> str:
        """
        Normaliza etiquetas de dimensão para facilitar correspondência.
        
        Args:
            label: Etiqueta a normalizar
            
        Returns:
            Etiqueta normalizada
        """
        if pd.isna(label) or label is None:
            return 'Total'
        
        str_label = str(label).strip()
        
        # Trata casos especiais antes dos padrões - mais abrangente
        special_cases = [
            '(em branco)', 'em branco', 'blank', '(blank)', 'vazio', '(vazio)', '',
            'total', 'todos', 'geral', 'sum', 'soma', 'all', 'conjunto',
            '(total)', '(todos)', '(geral)', 'totais'
        ]
        if str_label.lower() in special_cases:
            return 'Total'
        
        # NOVO: Aplica equivalências semânticas primeiro
        if str_label in self.semantic_equivalents:
            str_label = self.semantic_equivalents[str_label]
        
        # Aplica padrões de normalização
        for pattern, replacement in self.dimension_patterns.items():
            str_label = re.sub(pattern, replacement, str_label, flags=re.IGNORECASE)
        
        # Normalização adicional
        str_label = re.sub(r'\s+', ' ', str_label)  # Remove espaços múltiplos
        str_label = str_label.strip()
        
        return str_label
    
    def fuzzy_match_dimension(self, target: str, candidates: List[str], threshold: float = 0.8) -> Optional[str]:
        """
        Encontra correspondência difusa para uma dimensão.
        
        Args:
            target: Dimensão a procurar
            candidates: Lista de candidatos
            threshold: Limiar de similaridade (0-1)
            
        Returns:
            Melhor correspondência ou None se não encontrada
        """
        target_norm = self.normalize_dimension_label(target).lower()
        best_match = None
        best_score = 0
        
        for candidate in candidates:
            candidate_norm = self.normalize_dimension_label(candidate).lower()
            
            # Correspondência exata
            if target_norm == candidate_norm:
                return candidate
            
            # Correspondência difusa
            score = SequenceMatcher(None, target_norm, candidate_norm).ratio()
            if score > best_score and score >= threshold:
                best_score = score
                best_match = candidate
        
        return best_match
    
    def has_background_color(self, cell) -> bool:
        """
        Verifica se uma célula tem cor de fundo de forma robusta.
        
        Args:
            cell: Célula openpyxl
            
        Returns:
            True se tem cor de fundo, False caso contrário
        """
        try:
            if not cell.fill:
                return False
                
            # Verifica tipo de padrão
            if cell.fill.patternType and cell.fill.patternType != 'none':
                # Verifica cores RGB
                if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        # Cores que consideramos "sem fundo"
                        if rgb not in ['00000000', 'FFFFFFFF', None, 'FFFFFF', '000000']:
                            return True
                    
                    # Verifica theme colors
                    if hasattr(cell.fill.fgColor, 'theme') and cell.fill.fgColor.theme is not None:
                        return True
                        
                    # Verifica indexed colors
                    if hasattr(cell.fill.fgColor, 'indexed') and cell.fill.fgColor.indexed is not None:
                        # Índices 64 e 65 são geralmente automático/sem cor
                        return cell.fill.fgColor.indexed not in [64, 65]
                        
                return True  # Se tem padrão mas não conseguiu verificar cor, assume que tem cor
                
            return False
        except Exception as e:
            # Em caso de erro, assume que não tem cor (mais seguro)
            self.logger.debug(f"Erro ao verificar cor de fundo: {e}")
            return False
    
    def detect_header_by_color(self, cell, file_type='published') -> bool:
        """
        NOVO: Detecta se uma célula é cabeçalho baseado na cor de fundo.
        
        Args:
            cell: Célula openpyxl
            file_type: 'published' ou 'recreated' para aplicar regras específicas
            
        Returns:
            True se é cabeçalho baseado na cor
        """
        try:
            if not cell.fill or not cell.fill.patternType or cell.fill.patternType == 'none':
                return False
            
            # Estratégia específica para ficheiros publicados vs recriados
            if file_type == 'published':
                # Ficheiros publicados: azul escuro = cabeçalhos de coluna
                if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb).upper()
                    # Azuis escuros típicos em ficheiros publicados
                    dark_blues = ['FF366092', 'FF4F81BD', 'FF2F5597', 'FF1F497D']
                    return rgb in dark_blues
                    
                # Verifica theme colors para azuis
                if hasattr(cell.fill.fgColor, 'theme') and cell.fill.fgColor.theme is not None:
                    # Theme 1-6 são geralmente esquemas de azul
                    return cell.fill.fgColor.theme in [1, 2, 4, 5]
                    
            else:  # recreated files
                # Ficheiros recriados: azul claro = todos os cabeçalhos
                if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb).upper()
                    # Azuis claros típicos em ficheiros recriados
                    light_blues = ['FFCCE4F7', 'FFD9E2EC', 'FFDBE5F1', 'FFBFCDDB']
                    return rgb in light_blues
                    
                # Qualquer cor de fundo pode ser cabeçalho em recriados
                return self.has_background_color(cell)
            
            return False
            
        except Exception as e:
            self.logger.debug(f"Erro ao detectar cabeçalho por cor: {e}")
            return False
    
    def detect_hierarchy_by_spacing(self, ws, row: int, col: int, direction: str = 'row') -> int:
        """
        NOVO: Detecta nível hierárquico baseado no espaçamento visual.
        
        Args:
            ws: Worksheet
            row: Linha da célula
            col: Coluna da célula  
            direction: 'row' para hierarquia horizontal, 'column' para vertical
            
        Returns:
            Nível hierárquico (1=principal, 2=secundário, etc.)
        """
        try:
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                return 0
                
            cell_text = str(cell.value).strip()
            
            # Verifica indentação no texto
            if cell_text.startswith('  ') or cell_text.startswith('\t'):
                return 2  # Nível secundário (indentado)
            
            # Verifica se é sublinhado ou tem marcadores de hierarquia
            if any(marker in cell_text for marker in ['*', '-', '•', '►']):
                return 2
                
            # Verifica posição relativa e células mescladas
            if direction == 'column':
                # Para colunas, verifica se há célula mesclada acima
                for check_row in range(max(1, row-3), row):
                    check_cell = ws.cell(row=check_row, column=col)
                    if self.get_merged_cell_value(ws, check_row, col) and check_cell.value:
                        return 2  # É subcategoria de algo acima
                        
            elif direction == 'row':
                # Para linhas, verifica se há célula mesclada à esquerda
                for check_col in range(max(1, col-3), col):
                    check_cell = ws.cell(row=row, column=check_col)
                    if self.get_merged_cell_value(ws, row, check_col) and check_cell.value:
                        return 2  # É subcategoria de algo à esquerda
            
            return 1  # Nível principal por defeito
            
        except Exception as e:
            self.logger.debug(f"Erro ao detectar hierarquia por espaçamento: {e}")
            return 1
    
    def find_data_table(self, ws) -> Tuple[int, int, int, int]:
        """
        ATUALIZADO: Detecta a área da tabela distinguindo cabeçalhos de valores.
        
        Returns:
            Tupla (min_row, max_row, min_col, max_col) da área de dados
        """
        min_row, max_row = None, None
        min_col, max_col = None, None
        
        # Estratégia melhorada: encontra área com valores numéricos
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Verifica se é um valor numérico (dados reais)
                    numeric_value = self.normalize_value(cell.value)
                    if numeric_value is not None and not self._is_likely_header(cell):
                        # É uma célula de dados numéricos
                        if min_row is None or cell.row < min_row:
                            min_row = cell.row
                        if max_row is None or cell.row > max_row:
                            max_row = cell.row
                        if min_col is None or cell.column < min_col:
                            min_col = cell.column
                        if max_col is None or cell.column > max_col:
                            max_col = cell.column
        
        if min_row is None:
            # Fallback: usa área completa da planilha
            min_row, max_row = 1, ws.max_row
            min_col, max_col = 1, ws.max_column
        else:
            # Expande ligeiramente para incluir cabeçalhos próximos
            min_row = max(1, min_row - 5)
            min_col = max(1, min_col - 5)
        
        self.logger.debug(f"Área da tabela detectada: linhas {min_row}-{max_row}, colunas {min_col}-{max_col}")
        return min_row, max_row, min_col, max_col
    
    def _is_likely_header(self, cell) -> bool:
        """
        ULTRA-CONSERVADOR: Só identifica cabeçalhos óbvios, aceita todos os outros valores.
        O objectivo é capturar TODOS os valores numéricos para comparação.
        
        Args:
            cell: Célula a verificar
            
        Returns:
            True apenas se for OBVIAMENTE um cabeçalho
        """
        if not cell.value:
            return True
            
        cell_str = str(cell.value).strip()
        
        # APENAS cabeçalhos de texto óbvios
        obvious_headers = [
            'anos', 'unidade:', 'fonte:', 'nota:', 'continente',
            'açores', 'madeira', 'total geral', 'região autónoma'
        ]
        
        if any(header in cell_str.lower() for header in obvious_headers):
            return True
            
        # APENAS se for uma string muito específica de cabeçalho
        if isinstance(cell.value, str):
            # Strings com ":" são geralmente labels
            if ':' in cell_str:
                return True
            # Strings muito curtas como "I.10", "I.11" 
            if len(cell_str) <= 4 and any(c in cell_str for c in ['.', 'I', 'X', 'V']):
                return True
        
        # CRÍTICO: TODOS os valores numéricos são considerados DADOS (não cabeçalhos)
        numeric_value = self.normalize_value(cell.value)
        if numeric_value is not None:
            # EXCEÇÃO: Apenas anos óbvios são cabeçalhos
            if 1900 <= numeric_value <= 2030:
                return True
            # TODOS os outros números são DADOS
            else:
                return False
        
        # Por defeito, assume que é dados (não cabeçalho)
        return False
    
    def get_merged_cell_value(self, ws, row: int, col: int) -> Any:
        """
        Obtém o valor de uma célula considerando células mescladas.
        
        Args:
            ws: Worksheet openpyxl
            row: Número da linha
            col: Número da coluna
            
        Returns:
            Valor da célula ou da célula mesclada principal
        """
        cell = ws.cell(row=row, column=col)
        
        # Se a célula tem valor, retorna diretamente
        if cell.value is not None:
            return cell.value
        
        # Verifica se está numa área mesclada
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                # Retorna o valor da célula principal (top-left da área mesclada)
                main_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return main_cell.value
        
        return None
    
    def get_cell_dimensions(self, ws, data_row: int, data_col: int, 
                       table_bounds: Tuple[int, int, int, int]) -> Dict[str, Any]:
        """
        ATUALIZADO: Obtém as dimensões de uma célula usando estrutura hierárquica consistente.
        Retorna sempre: column_level_1, column_level_2, row_level_1, row_level_2
        
        Args:
            ws: Worksheet openpyxl
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            table_bounds: Limites da tabela (min_row, max_row, min_col, max_col)
            
        Returns:
            Dicionário com estrutura consistente de coordenadas
        """
        # Detecta tipo de ficheiro baseado no nome da workbook
        file_type = 'recreated' if 'result' in str(ws.parent.path).lower() else 'published'
        
        # Extrai coordenadas hierárquicas
        column_level_1, column_level_2 = self.extract_column_coordinates(ws, data_row, data_col, file_type)
        row_level_1, row_level_2 = self.extract_row_coordinates(ws, data_row, data_col, file_type)
        
        # Cria estrutura consistente de coordenadas
        coordinates = {}
        
        if column_level_1:
            coordinates["column_level_1"] = column_level_1
        if column_level_2:
            coordinates["column_level_2"] = column_level_2
        if row_level_1:
            coordinates["row_level_1"] = row_level_1  
        if row_level_2:
            coordinates["row_level_2"] = row_level_2
        
        # Debug logging melhorado
        coord_str = ", ".join([f"{k}:{v}" for k, v in coordinates.items()])
        self.logger.debug(f"[{file_type.upper()}] Célula ({data_row},{data_col}): {coord_str}")
        
        return coordinates
    
    def extract_column_coordinates(self, ws, data_row: int, data_col: int, file_type: str = 'published') -> Tuple[str, str]:
        """
        EQUILIBRADO: Extrai coordenadas de coluna sem ser excessivamente restritivo.
        Foca em cabeçalhos de texto e anos, mas não rejeita dados legítimos.
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            file_type: 'published' ou 'recreated'
            
        Returns:
            Tupla (column_level_1, column_level_2) - CABEÇALHOS EQUILIBRADOS
        """
        column_level_1 = ""
        column_level_2 = ""
        
        # Procura cabeçalhos de coluna caminhando para cima
        for search_row in range(data_row - 1, max(1, data_row - 10), -1):
            cell = ws.cell(row=search_row, column=data_col)
            cell_value = self.get_merged_cell_value(ws, search_row, data_col)
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                
                # Ignora células vazias ou cabeçalhos genéricos
                if not cell_str or cell_str.lower() in ['', 'anos', 'mes', 'mês', 'unidade: n.º', 'unidade']:
                    continue
                
                # Verifica se é um valor numérico
                numeric_value = self.normalize_value(cell_value)
                
                # EQUILIBRADO: Apenas rejeita se for claramente um dado estatístico grande
                if numeric_value is not None:
                    # Anos são válidos como coordenadas
                    if 1900 <= numeric_value <= 2030:
                        is_year = True
                        normalized_value = str(int(numeric_value))
                    # Valores muito grandes são provavelmente dados, não coordenadas
                    elif numeric_value > 100000:
                        self.logger.debug(f"[COORD] Ignorando valor estatístico grande: {cell_str}")
                        continue
                    else:
                        # Valores menores podem ser coordenadas válidas
                        normalized_value = self.normalize_dimension_label(cell_str)
                        is_year = False
                else:
                    # Valor de texto
                    normalized_value = self.normalize_dimension_label(cell_str)
                    is_year = False
                
                # Detecta se é cabeçalho por cor de fundo
                is_header_by_color = self.detect_header_by_color(cell, file_type)
                
                # Detecta nível hierárquico
                hierarchy_level = self.detect_hierarchy_by_spacing(ws, search_row, data_col, 'column')
                
                # Aceita se for ano, texto ou cabeçalho colorido
                is_text_header = isinstance(cell_value, str) or is_header_by_color
                
                if numeric_value is not None and (is_year or numeric_value <= 100000) or is_text_header:
                    if is_year:
                        if not column_level_1:
                            column_level_1 = normalized_value
                    elif is_header_by_color or hierarchy_level > 1:
                        if not column_level_2 and hierarchy_level == 2:
                            column_level_2 = normalized_value
                        elif not column_level_1 and hierarchy_level == 1:
                            column_level_1 = normalized_value
                    elif len(cell_str) > 0:
                        if not column_level_1:
                            column_level_1 = normalized_value
                
                # Para de procurar se encontrou ambos os níveis
                if column_level_1 and column_level_2:
                    break
        
        # Fallback limpo
        if not column_level_1:
            column_level_1 = f"Col_{data_col}"
            
        return (column_level_1, column_level_2)
    
    def extract_row_coordinates(self, ws, data_row: int, data_col: int, file_type: str = 'published') -> Tuple[str, str]:
        """
        EQUILIBRADO: Extrai coordenadas de linha de forma equilibrada.
        Foca em cabeçalhos de texto, mas não rejeita valores menores.
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            file_type: 'published' ou 'recreated'
            
        Returns:
            Tupla (row_level_1, row_level_2) - CABEÇALHOS EQUILIBRADOS
        """
        row_level_1 = ""
        row_level_2 = ""
        
        # Procura cabeçalhos de linha caminhando para a esquerda
        for search_col in range(data_col - 1, max(1, data_col - 10), -1):
            cell = ws.cell(row=data_row, column=search_col)
            cell_value = self.get_merged_cell_value(ws, data_row, search_col)
            
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                
                # Ignora células vazias ou cabeçalhos genéricos
                if not cell_str or cell_str.lower() in ['', 'mes', 'mês', 'anos']:
                    continue
                
                # Verifica se é um valor numérico
                numeric_value = self.normalize_value(cell_value)
                
                # EQUILIBRADO: Rejeita apenas valores claramente estatísticos ou anos
                if numeric_value is not None:
                    # Anos sempre rejeitados em coordenadas de linha
                    if 1900 <= numeric_value <= 2030:
                        self.logger.debug(f"[ROW_COORD] Ignorando ano: {cell_str}")
                        continue
                    # Valores estatísticos grandes também rejeitados
                    elif numeric_value > 10000:
                        self.logger.debug(f"[ROW_COORD] Ignorando valor estatístico: {cell_str}")
                        continue
                    # Valores pequenos podem ser códigos ou identificadores válidos
                    # else: continua o processamento
                
                # ACEITA texto ou números pequenos como possíveis coordenadas
                if isinstance(cell_value, str) or (numeric_value is not None and numeric_value <= 10000):
                    # Detecta se é cabeçalho por cor de fundo
                    is_header_by_color = self.detect_header_by_color(cell, file_type)
                    
                    # Detecta nível hierárquico
                    hierarchy_level = self.detect_hierarchy_by_spacing(ws, data_row, search_col, 'row')
                    
                    # Normaliza valor
                    normalized_value = self.normalize_dimension_label(cell_str)
                    
                    # Verifica se é mês (sempre level 2 ou special case)
                    is_month = any(month in cell_str.lower() for month in 
                                 ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                                  'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro',
                                  'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                                  'jul', 'ago', 'set', 'out', 'nov', 'dez', 'total'])
                    
                    if is_month:
                        if not row_level_2:
                            row_level_2 = normalized_value
                    elif is_header_by_color or hierarchy_level > 1:
                        if not row_level_2 and hierarchy_level == 2:
                            row_level_2 = normalized_value
                        elif not row_level_1 and hierarchy_level == 1:
                            row_level_1 = normalized_value
                    elif len(cell_str) > 0:
                        # Texto genérico - determina nível baseado na ordem
                        if not row_level_1:
                            row_level_1 = normalized_value
                        elif not row_level_2:
                            row_level_2 = normalized_value
                    
                    # Para de procurar se encontrou ambos os níveis
                    if row_level_1 and row_level_2:
                        break
        
        # Fallback limpo
        if not row_level_1:
            row_level_1 = f"Row_{data_row}"
            
        return (row_level_1, row_level_2)
    
    def get_displayed_value(self, cell) -> Optional[float]:
        """
        Obtém o valor apresentado de uma célula, usando abordagem mais inclusiva.
        
        Args:
            cell: Célula openpyxl
            
        Returns:
            Valor numérico conforme apresentado na célula, ou None se for cabeçalho
        """
        if cell.value is None:
            return None
        
        # Primeiro tenta normalizar qualquer valor (número ou texto)
        normalized_value = None
        
        if isinstance(cell.value, (int, float)):
            normalized_value = float(cell.value)
        else:
            # Tenta normalizar texto para número
            normalized_value = self.normalize_value(cell.value)
            
            # Se é texto e não se converte para número, verifica se é cabeçalho óbvio
            if normalized_value is None and isinstance(cell.value, str):
                str_value = str(cell.value).strip().lower()
                # Lista mais específica de padrões que SÃO cabeçalhos
                explicit_headers = [
                    'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                    'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro',
                    'january', 'february', 'march', 'april', 'may', 'june',
                    'july', 'august', 'september', 'october', 'november', 'december',
                    'indicador', 'unidade', 'fonte', 'nota', 'total', 'jan', 'fev', 'mar',
                    'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez',
                    'mês', 'mes', 'month', 'ano', 'anos', 'year', 'years'
                ]
                
                # Se contém cabeçalho explícito, não é valor
                if any(header == str_value or header in str_value.split() for header in explicit_headers):
                    return None
                
                # Se não conseguiu normalizar e não é cabeçalho explícito, também não é valor
                return None
        
        # Se conseguiu normalizar para número
        if normalized_value is not None:
            # FILTRO RIGOROSO: Anos típicos (1900-2030) NÃO são valores para comparar
            # São sempre dimensões, nunca dados estatísticos
            if (1900 <= normalized_value <= 2030 and 
                normalized_value == int(normalized_value)):
                # Anos são SEMPRE dimensões, nunca valores para comparar
                return None
            
            # Também rejeita outros valores pequenos que podem ser anos
            if (1800 <= normalized_value <= 2100 and 
                normalized_value == int(normalized_value) and
                normalized_value < 10000):
                # Estes são provavelmente anos ou identificadores, não dados estatísticos
                return None
            
            # Aceita todos os outros valores numéricos
            # Aplica formatação se necessário
            displayed_value = normalized_value
            
            if hasattr(cell, 'number_format') and cell.number_format and cell.number_format != 'General':
                # Para formatos com separadores de milhares ou decimais específicos
                if '0.0' in cell.number_format or '0.00' in cell.number_format:
                    # Arredonda conforme a formatação
                    try:
                        decimal_places = cell.number_format.count('0') - len(cell.number_format.split('.')[0].replace('#', '').replace('0', ''))
                        if decimal_places > 0:
                            displayed_value = round(float(normalized_value), decimal_places)
                    except:
                        pass  # Se der erro, usa valor original
            
            return displayed_value
        
        return None
    
    def detect_crosstab_structure(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        ULTRA-AGRESSIVO: Detecta TODOS os valores numéricos como dados.
        O objetivo é comparar TODOS os valores entre ficheiros.
        
        Args:
            file_path: Caminho do ficheiro Excel
            sheet_name: Nome da folha
            
        Returns:
            Dicionário com TODOS os valores numéricos encontrados
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                return {'error': f'Folha "{sheet_name}" não encontrada'}
            
            ws = wb[sheet_name]
            table_bounds = self.find_data_table(ws)
            
            # Detecta tipo de ficheiro para logging
            file_type = 'recreated' if 'result' in file_path.lower() else 'published'
            
            # Lista para armazenar TODOS os valores numéricos
            data_cells = []
            header_cells_found = 0
            total_cells_checked = 0
            
            min_row, max_row, min_col, max_col = table_bounds
            
            self.logger.info(f"[{file_type.upper()}] Analisando área {table_bounds} da folha '{sheet_name}'")
            
            # ESTRATÉGIA ULTRA-AGRESSIVA: Verifica cada célula
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_value = cell.value
                    total_cells_checked += 1
                    
                    if cell_value is not None:
                        # ACEITA QUALQUER valor numérico
                        normalized_value = self.normalize_value(cell_value)
                        
                        if normalized_value is not None:
                            # NOVA LÓGICA: Aceita TODOS os números, exceto cabeçalhos óbvios
                            is_obvious_header = self._is_likely_header(cell)
                            
                            if is_obvious_header:
                                header_cells_found += 1
                                self.logger.debug(f"[{file_type}_HEADER] ({row},{col}): {cell_value}")
                            else:
                                # ACEITA como dados - SEMPRE!
                                data_cells.append({
                                    'row': row,
                                    'col': col,
                                    'value': normalized_value,
                                    'cell': cell,
                                    'original_value': cell_value
                                })
                                
                                # Log primeiros 20 exemplos
                                if len(data_cells) <= 20:
                                    self.logger.info(f"[{file_type}_DATA] ({row},{col}): {normalized_value}")
                        else:
                            # Texto - pode ser cabeçalho
                            if isinstance(cell_value, str) and len(str(cell_value).strip()) > 0:
                                header_cells_found += 1
                                if header_cells_found <= 10:  # Log primeiros 10
                                    self.logger.debug(f"[{file_type}_TEXT] ({row},{col}): {cell_value}")
            
            # Estatísticas finais
            self.logger.info(f"[{file_type.upper()}] Folha '{sheet_name}': {len(data_cells)} dados, {header_cells_found} cabeçalhos, {total_cells_checked} total")
            
            # DEBUG CRÍTICO: Se não encontrou dados, mostra TUDO
            if len(data_cells) == 0:
                self.logger.error(f"[{file_type.upper()}] CRÍTICO: Nenhuma célula de dados encontrada!")
                self.logger.error(f"[{file_type.upper()}] Arquivo: {file_path}")
                self.logger.error(f"[{file_type.upper()}] Área analisada: {table_bounds}")
                
                # Mostra TODAS as células com valores para debug
                self.logger.error("=== DEBUG: TODAS as células com valores ===")
                debug_count = 0
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            norm_val = self.normalize_value(cell.value)
                            is_head = self._is_likely_header(cell)
                            has_color = self.has_background_color(cell)
                            self.logger.error(f"  ({row},{col}): '{cell.value}' -> norm:{norm_val}, header:{is_head}, color:{has_color}")
                            debug_count += 1
                            if debug_count >= 50:  # Limite para não sobrecarregar
                                self.logger.error("  ... (truncado após 50 células)")
                                break
                    if debug_count >= 50:
                        break
            
            return {
                'worksheet': ws,
                'table_bounds': table_bounds,
                'data_cells': data_cells,
                'total_data_cells': len(data_cells),
                'file_type': file_type
            }
            
        except Exception as e:
            self.logger.error(f"Erro ao detectar estrutura de {file_path}: {e}")
            return {'error': str(e)}
    
    def extract_crosstab_data(self, structure: Dict[str, Any]) -> Dict[Tuple, Any]:
        """
        ATUALIZADO: Extrai dados usando a nova estrutura hierárquica consistente.
        
        Args:
            structure: Estrutura detectada pela função detect_crosstab_structure
            
        Returns:
            Dicionário mapeando coordenadas hierárquicas para dados da célula
        """
        if 'error' in structure:
            return {}
        
        ws = structure['worksheet']
        table_bounds = structure['table_bounds']
        data_cells = structure['data_cells']
        data_map = {}
        
        # Detecta tipo de ficheiro
        file_type = 'recreated' if 'result' in str(ws.parent.path).lower() else 'published'
        
        # Processa cada célula de dados
        for data_cell in data_cells:
            row = data_cell['row']
            col = data_cell['col']
            value = data_cell['value']
            cell = data_cell['cell']
            
            # Obtém coordenadas hierárquicas consistentes
            coordinates_dict = self.get_cell_dimensions(ws, row, col, table_bounds)
            
            # Cria chave ordenada baseada na estrutura hierárquica
            ordered_coords = []
            
            # Ordem consistente: column_level_1, column_level_2, row_level_1, row_level_2
            for level in ['column_level_1', 'column_level_2', 'row_level_1', 'row_level_2']:
                if level in coordinates_dict:
                    ordered_coords.append(f"{level}:{coordinates_dict[level]}")
            
            # Se não há coordenadas hierárquicas, usa fallback
            if not ordered_coords:
                ordered_coords = [f"row_{row}", f"col_{col}"]
            
            coords_key = tuple(ordered_coords)
            
            # Debug específico para valores "Total"
            if any('Total' in str(coord) for coord in ordered_coords):
                self.logger.info(f"[{file_type.upper()} TOTAL] Célula ({row},{col}) valor={value}: {coords_key}")
            
            data_map[coords_key] = {
                'value': value,
                'row': row,
                'col': col,
                'cell': cell,
                'coordinates': coordinates_dict,
                'file_type': file_type
            }
        
        return data_map
    
    def compare_data_maps(self, published_map: Dict[Tuple, Any], 
                         recreated_map: Dict[Tuple, Any], 
                         sheet_name: str) -> List[Dict[str, Any]]:
        """
        ATUALIZADO: Compara mapas usando estrutura hierárquica consistente e equivalências semânticas.
        
        Args:
            published_map: Dados do ficheiro publicado
            recreated_map: Dados do ficheiro recriado
            sheet_name: Nome da folha
            
        Returns:
            Lista de discrepâncias encontradas
        """
        discrepancies = []
        
        # Verifica valores no ficheiro recriado
        for coords, recreated_data in recreated_map.items():
            recreated_value = recreated_data['value']
            recreated_row = recreated_data['row']
            recreated_col = recreated_data['col']
            recreated_cell = recreated_data['cell']
            
            # 1. CORRESPONDÊNCIA EXATA
            if coords in published_map:
                published_data = published_map[coords]
                published_value = published_data['value']
                
                # Compara valores com tolerância
                if abs(recreated_value - published_value) > self.numeric_tolerance:
                    discrepancies.append({
                        'sheet': sheet_name,
                        'coordinates': coords,
                        'recreated_value': recreated_value,
                        'published_value': published_value,
                        'difference': recreated_value - published_value,
                        'recreated_row': recreated_row,
                        'recreated_col': recreated_col,
                        'recreated_cell': recreated_cell,
                        'match_type': 'exact'
                    })
                continue  # Encontrou correspondência exata
            
            # 2. CORRESPONDÊNCIA COM EQUIVALÊNCIAS SEMÂNTICAS
            semantic_match_found = False
            for pub_coords, pub_data in published_map.items():
                if self._coords_semantically_equivalent(coords, pub_coords):
                    published_value = pub_data['value']
                    
                    if abs(recreated_value - published_value) > self.numeric_tolerance:
                        discrepancies.append({
                            'sheet': sheet_name,
                            'coordinates': coords,
                            'matched_coordinates': pub_coords,
                            'recreated_value': recreated_value,
                            'published_value': published_value,
                            'difference': recreated_value - published_value,
                            'recreated_row': recreated_row,
                            'recreated_col': recreated_col,
                            'recreated_cell': recreated_cell,
                            'match_type': 'semantic_exact'
                        })
                    
                    semantic_match_found = True
                    break
            
            if semantic_match_found:
                continue
            
            # 3. CORRESPONDÊNCIA DIFUSA AVANÇADA
            best_match = None
            best_match_coords = None
            best_score = 0
            
            for pub_coords, pub_data in published_map.items():
                score = self._calculate_coordinate_similarity(coords, pub_coords)
                
                if score > best_score and score >= 0.7:  # Threshold mais rigoroso
                    best_score = score
                    best_match = pub_data
                    best_match_coords = pub_coords
            
            if best_match:
                published_value = best_match['value']
                
                if abs(recreated_value - published_value) > self.numeric_tolerance:
                    discrepancies.append({
                        'sheet': sheet_name,
                        'coordinates': coords,
                        'matched_coordinates': best_match_coords,
                        'recreated_value': recreated_value,
                        'published_value': published_value,
                        'difference': recreated_value - published_value,
                        'recreated_row': recreated_row,
                        'recreated_col': recreated_col,
                        'recreated_cell': recreated_cell,
                        'match_type': 'fuzzy_advanced',
                        'match_score': best_score
                    })
            else:
                # 4. VALOR NÃO ENCONTRADO
                discrepancies.append({
                    'sheet': sheet_name,
                    'coordinates': coords,
                    'recreated_value': recreated_value,
                    'published_value': None,
                    'difference': None,
                    'recreated_row': recreated_row,
                    'recreated_col': recreated_col,
                    'recreated_cell': recreated_cell,
                    'match_type': 'not_found'
                })
        
        return discrepancies
    
    def _coords_semantically_equivalent(self, coords1: Tuple, coords2: Tuple) -> bool:
        """
        NOVO: Verifica se duas coordenadas são semanticamente equivalentes.
        
        Args:
            coords1: Primeira coordenada
            coords2: Segunda coordenada
            
        Returns:
            True se são semanticamente equivalentes
        """
        if len(coords1) != len(coords2):
            return False
        
        for coord1, coord2 in zip(coords1, coords2):
            # Extrai nível e valor de cada coordenada
            if ':' in str(coord1) and ':' in str(coord2):
                level1, value1 = str(coord1).split(':', 1)
                level2, value2 = str(coord2).split(':', 1)
                
                # Os níveis devem corresponder
                if level1 != level2:
                    return False
                
                # Verifica equivalência semântica dos valores
                if not self._values_semantically_equivalent(value1, value2):
                    return False
            else:
                # Correspondência direta para coordenadas simples
                if str(coord1) != str(coord2):
                    return False
        
        return True
    
    def _values_semantically_equivalent(self, value1: str, value2: str) -> bool:
        """
        NOVO: Verifica se dois valores são semanticamente equivalentes.
        
        Args:
            value1: Primeiro valor
            value2: Segundo valor
            
        Returns:
            True se são semanticamente equivalentes
        """
        # Normaliza valores primeiro
        norm_value1 = self.normalize_dimension_label(value1)
        norm_value2 = self.normalize_dimension_label(value2)
        
        # Correspondência direta
        if norm_value1 == norm_value2:
            return True
        
        # Verifica equivalências semânticas
        if norm_value1 in self.semantic_equivalents:
            return self.semantic_equivalents[norm_value1] == norm_value2
        
        if norm_value2 in self.semantic_equivalents:
            return self.semantic_equivalents[norm_value2] == norm_value1
        
        return False
    
    def _calculate_coordinate_similarity(self, coords1: Tuple, coords2: Tuple) -> float:
        """
        NOVO: Calcula similaridade entre coordenadas hierárquicas.
        
        Args:
            coords1: Primeira coordenada
            coords2: Segunda coordenada
            
        Returns:
            Score de similaridade (0-1)
        """
        if len(coords1) != len(coords2):
            # Penaliza diferenças de estrutura, mas permite alguma flexibilidade
            max_len = max(len(coords1), len(coords2))
            min_len = min(len(coords1), len(coords2))
            length_penalty = min_len / max_len
        else:
            length_penalty = 1.0
        
        matches = 0
        total_comparisons = min(len(coords1), len(coords2))
        
        for i in range(total_comparisons):
            coord1 = str(coords1[i])
            coord2 = str(coords2[i])
            
            # Correspondência exata
            if coord1 == coord2:
                matches += 1
            elif ':' in coord1 and ':' in coord2:
                # Compara níveis e valores separadamente
                level1, value1 = coord1.split(':', 1)
                level2, value2 = coord2.split(':', 1)
                
                if level1 == level2:  # Mesmo nível hierárquico
                    if self._values_semantically_equivalent(value1, value2):
                        matches += 1  # Equivalência semântica perfeita
                    else:
                        # Correspondência difusa de valores
                        fuzzy_score = SequenceMatcher(None, value1.lower(), value2.lower()).ratio()
                        if fuzzy_score >= 0.8:
                            matches += fuzzy_score  # Correspondência difusa ponderada
        
        base_score = matches / total_comparisons if total_comparisons > 0 else 0
        return base_score * length_penalty
    
    def compare_files(self, published_file: str, recreated_file: str, 
                     sheet_names: List[str]) -> Dict[str, Any]:
        """
        REDESENHADO: Compara ficheiros usando abordagem direcional simples.
        Preserva 100% das funcionalidades visuais existentes.
        
        Args:
            published_file: Caminho do ficheiro publicado
            recreated_file: Caminho do ficheiro recriado
            sheet_names: Lista de nomes das folhas a comparar
            
        Returns:
            Dicionário com resultados detalhados da comparação
        """
        self.logger.info("🚀 INICIANDO COMPARAÇÃO COM SISTEMA REDESENHADO")
        self.logger.info(f"📄 Publicado: {published_file}")
        self.logger.info(f"📄 Recriado: {recreated_file}")
        self.logger.info(f"📋 Folhas: {sheet_names}")
        
        results = {
            'published_file': published_file,
            'recreated_file': recreated_file,
            'sheet_results': {},
            'overall_stats': {
                'total_published_points': 0,
                'total_recreated_points': 0,
                'total_correct_matches': 0,
                'total_value_differences': 0,
                'total_missing_in_published': 0,
                'total_missing_in_recreated': 0,
                'overall_accuracy': 0.0
            }
        }
        
        for sheet_name in sheet_names:
            self.logger.info(f"\n📊 Processando folha: {sheet_name}")
            
            try:
                # NOVA ABORDAGEM: Extração simples directa
                self.logger.info(f"[{sheet_name}] Extraindo dados do ficheiro PUBLICADO...")
                published_points = self.extract_simple_data_points(published_file, sheet_name, 'published')
                
                self.logger.info(f"[{sheet_name}] Extraindo dados do ficheiro RECRIADO...")
                recreated_points = self.extract_simple_data_points(recreated_file, sheet_name, 'recreated')
                
                # NOVA ABORDAGEM: Comparação simples directa
                self.logger.info(f"[{sheet_name}] Comparando dados...")
                comparison_results = self.compare_simple_data(published_points, recreated_points, sheet_name)
                
                # Armazena resultados para esta folha
                results['sheet_results'][sheet_name] = comparison_results
                
                # Actualiza estatísticas globais
                stats = results['overall_stats']
                stats['total_published_points'] += len(published_points)
                stats['total_recreated_points'] += len(recreated_points)
                stats['total_correct_matches'] += len(comparison_results['correct_matches'])
                stats['total_value_differences'] += len(comparison_results['value_differences'])
                stats['total_missing_in_published'] += len(comparison_results['missing_in_published'])
                stats['total_missing_in_recreated'] += len(comparison_results['missing_in_recreated'])
                
                self.logger.info(f"[{sheet_name}] ✅ Concluído: {comparison_results['accuracy']:.2f}% de precisão")
                
            except Exception as e:
                self.logger.error(f"[{sheet_name}] ❌ Erro: {e}")
                results['sheet_results'][sheet_name] = {
                    'error': str(e),
                    'published_count': 0,
                    'recreated_count': 0,
                    'accuracy': 0.0,
                    'correct_matches': [],
                    'value_differences': [],
                    'missing_in_published': [],
                    'missing_in_recreated': []
                }
        
        # Calcula precisão geral
        total_recreated = results['overall_stats']['total_recreated_points']
        total_correct = results['overall_stats']['total_correct_matches']
        
        if total_recreated > 0:
            overall_accuracy = (total_correct / total_recreated) * 100
        else:
            overall_accuracy = 0.0
            
        results['overall_stats']['overall_accuracy'] = overall_accuracy
        
        # CRÍTICO: Adiciona chave 'summary' para compatibilidade (estava em falta!)
        results['summary'] = {
            'total_published_points': results['overall_stats']['total_published_points'],
            'total_recreated_points': results['overall_stats']['total_recreated_points'],
            'total_correct_matches': results['overall_stats']['total_correct_matches'],
            'total_discrepancies': results['overall_stats']['total_value_differences'] + results['overall_stats']['total_missing_in_published'],
            'accuracy_percentage': overall_accuracy
        }
        
        self.logger.info(f"\n🎯 RESULTADOS FINAIS:")
        self.logger.info(f"   📊 Total pontos publicados: {results['overall_stats']['total_published_points']}")
        self.logger.info(f"   📊 Total pontos recriados: {results['overall_stats']['total_recreated_points']}")
        self.logger.info(f"   ✅ Correspondências correctas: {results['overall_stats']['total_correct_matches']}")
        self.logger.info(f"   ❌ Diferenças de valor: {results['overall_stats']['total_value_differences']}")
        self.logger.info(f"   ❓ Ausentes no publicado: {results['overall_stats']['total_missing_in_published']}")
        self.logger.info(f"   ❓ Ausentes no recriado: {results['overall_stats']['total_missing_in_recreated']}")
        self.logger.info(f"   🎯 PRECISÃO GERAL: {overall_accuracy:.2f}%")
        
        return results
    
    def copy_worksheet_with_formatting(self, source_ws, target_wb, target_name: str):
        """
        Copia uma folha preservando toda a formatação original.
        
        Args:
            source_ws: Worksheet origem
            target_wb: Workbook destino
            target_name: Nome da folha destino
            
        Returns:
            Worksheet copiada
        """
        target_ws = target_wb.create_sheet(target_name)
        
        # Copia dimensões das colunas
        for col_letter in source_ws.column_dimensions:
            if source_ws.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
        
        # Copia dimensões das linhas
        for row_num in source_ws.row_dimensions:
            if source_ws.row_dimensions[row_num].height:
                target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height
        
        # Copia células e formatação
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                
                # Copia valor
                target_cell.value = cell.value
                
                # Copia formatação
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)
        
        # Copia células mescladas
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))
        
        return target_ws
    
    def create_highlighted_report_sheet(self, source_ws, target_wb, sheet_name: str, 
                                       comparison_results: Dict[str, Any]):
        """
        REDESENHADO: Cria folha de relatório com sistema de highlighting de duas cores.
        
        Args:
            source_ws: Folha de origem (recriada)
            target_wb: Workbook de destino
            sheet_name: Nome da folha 
            comparison_results: Resultados da comparação simples
        """
        # Copia folha com formatação
        target_ws = self.copy_worksheet_with_formatting(source_ws, target_wb, f"{sheet_name}_comparison")
        
        # NOVO: Aplica highlighting de duas cores com comentários detalhados
        correct_count, incorrect_count = self.apply_enhanced_highlighting(target_ws, sheet_name, comparison_results)
        
        # Ajusta largura das colunas para melhor visualização
        for col in ['A', 'B', 'C', 'D', 'E']:
            target_ws.column_dimensions[col].width = 15
            
        total_highlighted = correct_count + incorrect_count
        self.logger.info(f"Folha de comparação criada: {sheet_name}_comparison com {total_highlighted} células destacadas ({correct_count} verdes, {incorrect_count} amarelas)")
    
    def _create_enhanced_comment(self, discrepancy: Dict[str, Any]) -> str:
        """
        NOVO: Cria comentário detalhado para células com discrepâncias.
        
        Args:
            discrepancy: Dicionário com informações da discrepância
            
        Returns:
            Texto do comentário formatado
        """
        comment_lines = ["=== DISCREPÂNCIA DETECTADA ==="]
        
        # Valores
        recreated_val = discrepancy.get('recreated_value', 'N/A')
        published_val = discrepancy.get('published_value', 'N/A')
        
        comment_lines.append(f"Valor Recriado: {recreated_val}")
        
        if published_val != 'N/A' and published_val is not None:
            comment_lines.append(f"Valor Publicado: {published_val}")
            
            # Calcula diferença
            try:
                diff = float(recreated_val) - float(published_val)
                if abs(diff) > 0.001:  # Evita diferenças de arredondamento
                    comment_lines.append(f"Diferença: {diff:+.1f}")
            except:
                comment_lines.append("Diferença: N/A")
        else:
            comment_lines.append("Valor Publicado: NÃO ENCONTRADO")
        
        # Coordenadas limpas
        coords = discrepancy.get('coordinates', ())
        if coords:
            clean_coords = self._format_coordinates_for_display(coords)
            comment_lines.append(f"Posição: {clean_coords}")
        
        # Tipo de correspondência
        match_type = discrepancy.get('match_type', 'unknown')
        type_descriptions = {
            'exact': 'Coordenadas exatas',
            'semantic_exact': 'Equivalência semântica',
            'fuzzy_advanced': 'Correspondência difusa',
            'not_found': 'Valor único'
        }
        comment_lines.append(f"Tipo: {type_descriptions.get(match_type, match_type)}")
        
        # Score para correspondências difusas
        if 'match_score' in discrepancy:
            score = discrepancy['match_score']
            comment_lines.append(f"Similaridade: {score:.1%}")
        
        return "\n".join(comment_lines)
    
    def _format_coordinates_for_display(self, coords: tuple) -> str:
        """
        NOVO: Formata coordenadas para exibição limpa em comentários.
        
        Args:
            coords: Tupla de coordenadas
            
        Returns:
            String formatada para exibição
        """
        if not coords:
            return "N/A"
        
        # Extrai valores limpos das coordenadas
        display_parts = []
        
        for coord in coords:
            coord_str = str(coord)
            if ':' in coord_str:
                level, value = coord_str.split(':', 1)
                
                # Simplifica nomes de níveis para exibição
                level_map = {
                    'column_level_1': 'Col',
                    'column_level_2': 'SubCol', 
                    'row_level_1': 'Linha',
                    'row_level_2': 'SubLinha'
                }
                
                display_level = level_map.get(level, level)
                display_parts.append(f"{display_level}: {value}")
            else:
                display_parts.append(coord_str)
        
        return " | ".join(display_parts)
    
    def generate_comparison_report(self, results: Dict[str, Any], output_dir: str = "result/comparison") -> str:
        """
        Gera relatório Excel visual detalhado da comparação com destaques.
        
        Args:
            results: Resultados da comparação
            output_dir: Diretório de saída
            
        Returns:
            Caminho do ficheiro de relatório gerado
        """
        ensure_directory_exists(output_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = os.path.join(output_dir, f"visual_comparison_report_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Remove folha padrão
        wb.remove(wb.active)
        
        # Folha de resumo geral
        summary_ws = wb.create_sheet("Resumo_Geral")
        self._create_summary_sheet(summary_ws, results)
        
        # NOVO: Cria folhas visuais usando nova estrutura
        for sheet_name, sheet_results in results['sheet_results'].items():
            if 'error' not in sheet_results:
                try:
                    # Carrega folha do ficheiro recriado para usar como base
                    wb_recreated = openpyxl.load_workbook(results['recreated_file'], data_only=True)
                    if sheet_name in wb_recreated.sheetnames:
                        source_ws = wb_recreated[sheet_name]
                        
                        # Cria folha visual com highlighting de duas cores
                        self.create_highlighted_report_sheet(
                            source_ws, wb, sheet_name, sheet_results
                        )
                    
                    wb_recreated.close()
                    
                except Exception as e:
                    self.logger.error(f"Erro ao criar folha visual para {sheet_name}: {e}")
        
        # NOVO: Folha de discrepâncias detalhadas com nova estrutura
        for sheet_name, sheet_results in results['sheet_results'].items():
            if 'error' not in sheet_results and (sheet_results.get('value_differences') or sheet_results.get('missing_in_published')):
                disc_ws = wb.create_sheet(f"Detalhes_{sheet_name}"[:31])
                self._create_discrepancy_sheet(disc_ws, sheet_results, sheet_name)
        
        # Folha de detalhes técnicos
        details_ws = wb.create_sheet("Info_Tecnica")
        self._create_technical_details_sheet(details_ws, results)
        
        # Limpa recursos das estruturas
        for sheet_results in results['sheet_results'].values():
            if 'pub_structure' in sheet_results and 'workbook' in sheet_results['pub_structure']:
                sheet_results['pub_structure']['workbook'].close()
            if 'rec_structure' in sheet_results and 'workbook' in sheet_results['rec_structure']:
                sheet_results['rec_structure']['workbook'].close()
        
        wb.save(report_file)
        self.logger.info(f"Relatório visual gerado: {report_file}")
        
        return report_file
    
    def _create_summary_sheet(self, ws, results: Dict[str, Any]):
        """Cria folha de resumo do relatório."""
        # Cabeçalho
        ws['A1'] = "RELATÓRIO DE COMPARAÇÃO DE DADOS"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        ws[f'A{row}'] = "Ficheiro Publicado:"
        ws[f'B{row}'] = results['published_file']
        row += 1
        
        ws[f'A{row}'] = "Ficheiro Recriado:"
        ws[f'B{row}'] = results['recreated_file']
        row += 1
        
        ws[f'A{row}'] = "Data da Comparação:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 2
        
        # Estatísticas gerais
        ws[f'A{row}'] = "ESTATÍSTICAS GERAIS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        stats = results['overall_stats']
        ws[f'A{row}'] = "Total de pontos publicados:"
        ws[f'B{row}'] = stats['total_published_points']
        row += 1
        
        ws[f'A{row}'] = "Total de pontos recriados:"
        ws[f'B{row}'] = stats['total_recreated_points']
        row += 1
        
        ws[f'A{row}'] = "Correspondências correctas:"
        ws[f'B{row}'] = stats['total_correct_matches']
        row += 1
        
        ws[f'A{row}'] = "Diferenças de valor:"
        ws[f'B{row}'] = stats['total_value_differences']
        row += 1
        
        ws[f'A{row}'] = "Total de discrepâncias:"
        ws[f'B{row}'] = stats['total_missing_in_published'] + stats['total_missing_in_recreated']
        row += 1
        
        ws[f'A{row}'] = "Precisão:"
        ws[f'B{row}'] = f"{stats['overall_accuracy']:.2f}%"
        row += 2
        
        # Resumo por folha
        ws[f'A{row}'] = "RESUMO POR FOLHA"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Folha"
        ws[f'B{row}'] = "Pontos Publicados"
        ws[f'C{row}'] = "Pontos Recriados"
        ws[f'D{row}'] = "Discrepâncias"
        ws[f'E{row}'] = "Precisão (%)"
        
        # Formatação cabeçalho
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        
        for sheet_name, sheet_results in results['sheet_results'].items():
            if 'error' not in sheet_results:
                pub_points = sheet_results.get('published_data_points', 0)
                rec_points = sheet_results.get('recreated_data_points', 0)
                discrepancies = sheet_results.get('discrepancy_count', 0)
                accuracy = sheet_results.get('accuracy', 0.0)
                
                ws[f'A{row}'] = sheet_name
                ws[f'B{row}'] = pub_points
                ws[f'C{row}'] = rec_points
                ws[f'D{row}'] = discrepancies
                ws[f'E{row}'] = f"{accuracy:.2f}%"
                row += 1
        
        # Ajusta largura das colunas
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
    
    def _create_discrepancy_sheet(self, ws, sheet_results: Dict[str, Any], sheet_name: str):
        """NOVO: Cria folha de discrepâncias com nova estrutura de dados."""
        ws['A1'] = f"DISCREPÂNCIAS - {sheet_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Combina diferenças de valor e valores ausentes
        all_discrepancies = []
        
        # Diferenças de valor
        for diff in sheet_results.get('value_differences', []):
            all_discrepancies.append({
                'coordinates': f"({diff['coordinates'][0]}, {diff['coordinates'][1]})",
                'recreated_value': diff['recreated_value'],
                'published_value': diff['published_value'],
                'difference': diff['difference'],
                'type': 'Diferença de Valor'
            })
        
        # Valores ausentes no publicado
        for missing in sheet_results.get('missing_in_published', []):
            all_discrepancies.append({
                'coordinates': f"({missing['coordinates'][0]}, {missing['coordinates'][1]})",
                'recreated_value': missing['recreated_value'],
                'published_value': 'NÃO ENCONTRADO',
                'difference': 'N/A',
                'type': 'Ausente no Publicado'
            })
        
        if not all_discrepancies:
            ws['A3'] = "Nenhuma discrepância encontrada nesta folha."
            return
        
        # Cabeçalhos
        headers = ['Coordenadas', 'Valor Recriado', 'Valor Publicado', 'Diferença', 'Tipo']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados das discrepâncias
        for i, disc in enumerate(all_discrepancies, 4):
            ws.cell(row=i, column=1, value=disc['coordinates'])
            ws.cell(row=i, column=2, value=disc['recreated_value'])
            ws.cell(row=i, column=3, value=disc['published_value'])
            ws.cell(row=i, column=4, value=disc['difference'])
            ws.cell(row=i, column=5, value=disc['type'])
            
            # Destaca discrepâncias significativas
            if disc['difference'] != 'N/A' and isinstance(disc['difference'], (int, float)):
                if abs(disc['difference']) > self.numeric_tolerance * 10:
                    for col in range(1, 6):
                        ws.cell(row=i, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Ajusta largura das colunas
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 25
    
    def _create_technical_details_sheet(self, ws, results: Dict[str, Any]):
        """Cria folha com detalhes técnicos da comparação."""
        ws['A1'] = "DETALHES TÉCNICOS DA COMPARAÇÃO"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Parâmetros de comparação
        ws[f'A{row}'] = "PARÂMETROS DE COMPARAÇÃO"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Tolerância numérica:"
        ws[f'B{row}'] = self.numeric_tolerance
        row += 1
        
        ws[f'A{row}'] = "Limiar correspondência difusa:"
        ws[f'B{row}'] = "0.8"
        row += 2
        
        # Informações das folhas
        ws[f'A{row}'] = "INFORMAÇÕES DAS FOLHAS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for sheet_name, sheet_results in results['sheet_results'].items():
            ws[f'A{row}'] = f"Folha: {sheet_name}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if 'error' in sheet_results:
                ws[f'B{row}'] = f"Erro: {sheet_results['error']}"
                row += 1
            else:
                # CORRIGIDO: Usa as chaves correctas da nova estrutura
                correct_matches = len(sheet_results.get('correct_matches', []))
                value_diffs = len(sheet_results.get('value_differences', []))
                missing_pub = len(sheet_results.get('missing_in_published', []))
                total_discrepancies = value_diffs + missing_pub
                
                ws[f'B{row}'] = f"Correspondências correctas: {correct_matches}"
                row += 1
                ws[f'B{row}'] = f"Total de discrepâncias: {total_discrepancies}"
                row += 1
                ws[f'B{row}'] = f"  • Valores diferentes: {value_diffs}"
                row += 1
                ws[f'B{row}'] = f"  • Valores em falta: {missing_pub}"
                row += 1
            row += 1
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def find_row_header(self, ws, data_row: int, data_col: int) -> str:
        """
        NOVO: Busca ESPECÍFICA POR TIPO DE FICHEIRO de cabeçalhos de linha REAIS.
        
        PROBLEM IDENTIFIED: Published vs Recreated files have DIFFERENT LAYOUTS!
        
        ESTRATÉGIAS POR TIPO:
        - PUBLISHED: Estrutura específica com possível formato diferente  
        - RECREATED: Busca padrão que já está a funcionar
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            
        Returns:
            String com o cabeçalho da linha encontrado
        """
        
        # CRÍTICO: Detecta tipo de ficheiro baseado no caminho
        file_type = 'recreated' if 'result' in str(ws.parent.path).lower() else 'published'
        self.logger.debug(f"🔍 BUSCA ROW HEADER para {file_type.upper()} - linha {data_row}")
        
        if file_type == 'published':
            return self._find_row_header_published(ws, data_row, data_col)
        else:
            return self._find_row_header_recreated(ws, data_row, data_col)
    
    def _find_row_header_published(self, ws, data_row: int, data_col: int) -> str:
        """
        ESPECÍFICO: Busca row headers em ficheiros PUBLICADOS.
        
        Ficheiros publicados podem ter estrutura diferente:
        - Cabeçalhos numa posição específica
        - Layout de tabela diferente
        - Possível presença de células mescladas
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados  
            data_col: Coluna da célula de dados
            
        Returns:
            String com o cabeçalho encontrado
        """
        self.logger.debug(f"🔍 PUBLISHED: Busca específica para linha {data_row}")
        
        # ESTRATÉGIA A: Busca na coluna 1 (primeira coluna) - layout típico de publicados
        if ws.max_column >= 1:
            cell_col1 = ws.cell(row=data_row, column=1)
            if cell_col1.value and isinstance(cell_col1.value, str):
                clean_value = str(cell_col1.value).strip()
                if self._is_valid_row_header(clean_value):
                    self.logger.info(f"✅ PUBLISHED A: Row header na col 1: '{clean_value}' em ({data_row},1)")
                    return clean_value
        
        # ESTRATÉGIA B: Busca em colunas específicas típicas de publicados (2-5)
        for col in [2, 3, 4, 5]:
            if col <= ws.max_column:
                cell = ws.cell(row=data_row, column=col)
                if cell.value and isinstance(cell.value, str):
                    clean_value = str(cell.value).strip()
                    if self._is_valid_row_header(clean_value):
                        self.logger.info(f"✅ PUBLISHED B: Row header na col {col}: '{clean_value}' em ({data_row},{col})")
                        return clean_value
        
        # ESTRATÉGIA C: Busca em células mescladas em linha específica
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= data_row <= merged_range.max_row:
                merged_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                if merged_cell.value and isinstance(merged_cell.value, str):
                    clean_value = str(merged_cell.value).strip()
                    if self._is_valid_row_header(clean_value):
                        self.logger.info(f"✅ PUBLISHED C: Row header mesclado: '{clean_value}'")
                        return clean_value
        
        # ESTRATÉGIA D: Busca horizontal expandida (ESQUERDA) com mais alcance
        max_search_cols = min(data_col - 1, 15)  # Busca até 15 colunas
        
        for col in range(data_col - 1, max(0, data_col - max_search_cols), -1):
            cell = ws.cell(row=data_row, column=col)
            if cell.value and isinstance(cell.value, str):
                clean_value = str(cell.value).strip()
                if self._is_valid_row_header(clean_value):
                    self.logger.info(f"✅ PUBLISHED D: Row header horizontal: '{clean_value}' em ({data_row},{col})")
                    return clean_value
        
        # ESTRATÉGIA E: Busca vertical específica para published
        for offset in range(1, 8):  # Busca 7 linhas acima e abaixo
            for direction in [-1, 1]:
                target_row = data_row + (offset * direction)
                if 1 <= target_row <= ws.max_row:
                    # Verifica colunas específicas típicas
                    for check_col in [1, 2, 3, 4, 5]:
                        if check_col <= ws.max_column:
                            cell = ws.cell(row=target_row, column=check_col)
                            if cell.value and isinstance(cell.value, str):
                                clean_value = str(cell.value).strip()
                                if self._is_valid_row_header(clean_value):
                                    self.logger.info(f"✅ PUBLISHED E: Row header vertical: '{clean_value}' em ({target_row},{check_col})")
                                    return clean_value
        
        # FALLBACK: Todas as estratégias falharam
        fallback = f"Row_{data_row}"
        self.logger.warning(f"⚠️ PUBLISHED: Todas estratégias falharam para linha {data_row}, usando fallback: {fallback}")
        return fallback
    
    def _find_row_header_recreated(self, ws, data_row: int, data_col: int) -> str:
        """
        ESPECÍFICO: Busca row headers em ficheiros RECRIADOS.
        
        Esta lógica JÁ ESTÁ A FUNCIONAR conforme os logs mostram.
        Mantenho a estratégia original com pequenos ajustes.
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            
        Returns:
            String com o cabeçalho encontrado
        """
        self.logger.debug(f"🔍 RECREATED: Busca original para linha {data_row}")
        
        # ESTRATÉGIA 1: Busca horizontal (ESQUERDA) - esta está a funcionar
        max_search_cols = min(data_col - 1, int(ws.max_column * 0.8))
        
        for col in range(data_col - 1, max(0, data_col - max_search_cols), -1):
            cell = ws.cell(row=data_row, column=col)
            if cell.value and isinstance(cell.value, str):
                clean_value = str(cell.value).strip()
                if self._is_valid_row_header(clean_value):
                    self.logger.info(f"✅ RECREATED 1: Row header horizontal: '{clean_value}' em ({data_row},{col})")
                    return clean_value
        
        # ESTRATÉGIA 2: Busca vertical - mantém lógica original
        for offset in range(1, 11):
            for direction in [-1, 1]:
                target_row = data_row + (offset * direction)
                if 1 <= target_row <= ws.max_row:
                    for col_offset in range(-5, 6):
                        target_col = data_col + col_offset
                        if 1 <= target_col <= ws.max_column:
                            cell = ws.cell(row=target_row, column=target_col)
                            if cell.value and isinstance(cell.value, str):
                                clean_value = str(cell.value).strip()
                                if self._is_valid_row_header(clean_value):
                                    self.logger.info(f"✅ RECREATED 2: Row header vertical: '{clean_value}' em ({target_row},{target_col})")
                                    return clean_value
        
        # ESTRATÉGIA 3: Células mescladas
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= data_row <= merged_range.max_row:
                merged_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                if merged_cell.value and isinstance(merged_cell.value, str):
                    clean_value = str(merged_cell.value).strip()
                    if self._is_valid_row_header(clean_value):
                        self.logger.info(f"✅ RECREATED 3: Row header mesclado: '{clean_value}'")
                        return clean_value
        
        # ESTRATÉGIA 4: Padrões específicos
        target_patterns = ['Total', 'Hotéis', 'Hotelaria', 'Pensões', 'Motéis', 
                          'Alojamento', 'Estabelecimentos', 'Turismo', 'Abril', 
                          'Maio', 'Junho', 'Julho', 'Agosto', 'Janeiro', 'Fevereiro',
                          'Março', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=data_row, column=col)
            if cell.value and isinstance(cell.value, str):
                cell_text = str(cell.value).strip()
                for pattern in target_patterns:
                    if pattern.lower() in cell_text.lower():
                        self.logger.info(f"✅ RECREATED 4: Row header padrão '{pattern}': '{cell_text}' em ({data_row},{col})")
                        return cell_text
        
        # FALLBACK
        fallback = f"Row_{data_row}"
        self.logger.warning(f"⚠️ RECREATED: Todas estratégias falharam para linha {data_row}, usando fallback: {fallback}")
        return fallback
    
    def _is_valid_row_header(self, header_text: str) -> bool:
        """
        AUXILIAR: Valida se um texto é um cabeçalho de linha válido.
        
        Args:
            header_text: Texto a validar
            
        Returns:
            True se é um cabeçalho válido
        """
        if not header_text or len(header_text.strip()) < 2:
            return False
        
        clean_value = header_text.strip()
        
        # REJEITA placeholders
        invalid_headers = {'-', '.', '•', '*', '**', '***', '****', 
                         '_', '()', '( )', ':', ';', ',', '|', 
                         '/', '\\', '+', '=', '#', '@', '!', '?'}
        
        if clean_value in invalid_headers:
            return False
        
        # REJEITA números grandes (estatísticos)
        try:
            numeric_test = float(clean_value.replace(',', '.'))
            if numeric_test > 1000:  # Provavelmente dados estatísticos
                return False
        except ValueError:
            pass  # É texto - continua validação
        
        # ACEITA se contém padrões válidos
        valid_patterns = [
            'hotel', 'hotelaria', 'alojamento', 'total', 'estabelecimento',
            'pensão', 'motel', 'turismo', 'branco', 'em branco', 'geral',
            'continente', 'açores', 'madeira', 'região', 'nacional',
            'abril', 'maio', 'junho', 'julho', 'agosto', 'janeiro',
            'fevereiro', 'março', 'setembro', 'outubro', 'novembro', 'dezembro',
            'categoria', 'tipo', 'classe', 'star', 'estrela', 'norte',
            'centro', 'sul', 'lisboa', 'porto'
        ]
        
        clean_lower = clean_value.lower()
        
        # ACEITA se tem padrão válido OU é texto longo OU contém caracteres portugueses
        has_pattern = any(pattern in clean_lower for pattern in valid_patterns)
        is_long_text = len(clean_value) >= 5
        has_portuguese = any(c in clean_lower for c in ['ã', 'ç', 'é', 'í', 'ó', 'ú', 'â', 'ê', 'ô'])
        
        return has_pattern or is_long_text or has_portuguese
    
    def find_column_header(self, ws, data_row: int, data_col: int) -> str:
        """
        REESCRITO: Busca específica por tipo de ficheiro para column headers.
        
        PROBLEMA CORRIGIDO: Detecção mais robusta de tipo de ficheiro
        - Verifica se há dados em posições típicas de ficheiros published
        - Não depende apenas da linha de dados
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            
        Returns:
            String com o cabeçalho da coluna encontrado
        """
        
        # MELHORADA: Detecção mais robusta de tipo de ficheiro
        is_published_file = self._detect_published_file_type(ws, data_row, data_col)
        
        if is_published_file:
            # ESTRATÉGIA PUBLISHED: Busca especificamente em linhas onde anos costumam estar
            self.logger.debug(f"🔍 PUBLISHED: Busca column header para coluna {data_col}")
            
            # ESTRATÉGIA A: Busca nas linhas 1-6 primeiro (headers muito próximos ao topo)
            for check_row in [1, 2, 3, 4, 5, 6]:
                if check_row <= ws.max_row:
                    cell = ws.cell(row=check_row, column=data_col)
                    if cell.value and self.is_year(cell.value):
                        year_str = str(int(float(cell.value)))
                        self.logger.debug(f"✅ PUBLISHED A: Column header linha {check_row}: '{year_str}'")
                        return year_str
            
            # ESTRATÉGIA B: Busca na linha 16 (típica para published)
            if ws.max_row >= 16:
                cell_16 = ws.cell(row=16, column=data_col)
                if cell_16.value and self.is_year(cell_16.value):
                    year_str = str(int(float(cell_16.value)))
                    self.logger.debug(f"✅ PUBLISHED B: Column header linha 16: '{year_str}'")
                    return year_str
            
            # ESTRATÉGIA C: Busca na linha 17 (alternativa para published)
            if ws.max_row >= 17:
                cell_17 = ws.cell(row=17, column=data_col)
                if cell_17.value and self.is_year(cell_17.value):
                    year_str = str(int(float(cell_17.value)))
                    self.logger.debug(f"✅ PUBLISHED C: Column header linha 17: '{year_str}'")
                    return year_str
            
            # ESTRATÉGIA D: Busca nas linhas 14-20 (área típica de headers published)
            for check_row in [14, 15, 18, 19, 20]:
                if check_row <= ws.max_row:
                    cell = ws.cell(row=check_row, column=data_col)
                    if cell.value and self.is_year(cell.value):
                        year_str = str(int(float(cell.value)))
                        self.logger.debug(f"✅ PUBLISHED D: Column header linha {check_row}: '{year_str}'")
                        return year_str
        
        else:
            # ESTRATÉGIA RECREATED: Busca normal (está a funcionar)
            self.logger.debug(f"🔍 RECREATED: Busca column header para coluna {data_col}")
            
        # ESTRATÉGIA GERAL: Busca expandida para ambos os tipos
        max_search_rows = min(data_row - 1, 30)
        
        for row in range(data_row - 1, max(0, data_row - max_search_rows), -1):
            cell = ws.cell(row=row, column=data_col)
            cell_value = cell.value
            
            if cell_value:
                if isinstance(cell_value, str):
                    clean_value = str(cell_value).strip()
                    
                    # CRÍTICO: Lista de placeholders a rejeitar
                    invalid_headers = {'-', '.', '•', '*', '**', '***', '****', 
                                     '_', '()', '( )', ':', ';', ',', '|', 
                                     '/', '\\', '+', '=', '#', '@', '!', '?'}
                    
                    if clean_value in invalid_headers:
                        self.logger.debug(f"REJEITADO placeholder/símbolo '{clean_value}' como column header em ({row},{data_col})")
                        continue
                        
                    if len(clean_value) < 2:
                        self.logger.debug(f"REJEITADO texto muito curto '{clean_value}' como column header em ({row},{data_col})")
                        continue
                    
                    # Verifica se é um ano como string
                    try:
                        year_test = float(clean_value.replace(',', '.'))
                        if 1900 <= year_test <= 2030:
                            self.logger.debug(f"Column header (ano como texto) encontrado: '{clean_value}' em ({row},{data_col})")
                            return str(int(year_test))
                        elif year_test > 2030:
                            self.logger.debug(f"REJEITADO valor estatístico '{clean_value}' como column header em ({row},{data_col})")
                            continue
                    except ValueError:
                        # Texto legítimo
                        valid_column_patterns = [
                            'unidade', 'total', 'ano', 'mês', 'período',
                            'trimestre', 'semestre', 'fonte', 'nota'
                        ]
                        
                        clean_lower = clean_value.lower()
                        is_valid_header = any(pattern in clean_lower for pattern in valid_column_patterns)
                        
                        if is_valid_header or len(clean_value) >= 4:
                            self.logger.debug(f"Column header (texto legítimo) encontrado: '{clean_value}' em ({row},{data_col})")
                            return clean_value
                        else:
                            self.logger.debug(f"REJEITADO texto sem padrão válido '{clean_value}' como column header em ({row},{data_col})")
                            continue
                            
                elif isinstance(cell_value, (int, float)):
                    if 1900 <= cell_value <= 2030:
                        year_str = str(int(cell_value))
                        self.logger.debug(f"Column header (ano) encontrado: '{year_str}' em ({row},{data_col})")
                        return year_str
                    else:
                        self.logger.debug(f"REJEITADO valor '{cell_value}' como column header (fora do intervalo de anos) em ({row},{data_col})")
                        continue
        
        # FALLBACK
        fallback = f"Col_{data_col}"
        file_type = "PUBLISHED" if is_published_file else "RECREATED"
        self.logger.warning(f"⚠️ {file_type}: FALLBACK usado para coluna {data_col} após buscar {max_search_rows} linhas!")
        return fallback

    def _detect_published_file_type(self, ws, data_row: int, data_col: int) -> bool:
        """
        NOVO: Detecta mais robustamente se é um ficheiro published.
        
        Verifica múltiplos sinais:
        1. Posição de dados (linha 5 é published, linha 17+ é recreated)
        2. Presença de padrões típicos de published files
        3. Estrutura geral da planilha
        
        Args:
            ws: Worksheet
            data_row: Linha da célula de dados
            data_col: Coluna da célula de dados
            
        Returns:
            True se parecer ser um ficheiro published
        """
        # SINAL 1: Dados em linhas muito baixas (1-10) = Published
        if data_row <= 10:
            self.logger.debug(f"PUBLISHED detectado: dados na linha {data_row} (muito baixa)")
            return True
            
        # SINAL 2: Dados em linhas médias mas com padrões published
        if 10 < data_row < 20:
            # Verifica se há anos nas linhas 1-6 (típico de published)
            for check_row in [1, 2, 3, 4, 5, 6]:
                if check_row <= ws.max_row:
                    cell = ws.cell(row=check_row, column=data_col)
                    if cell.value and self.is_year(cell.value):
                        self.logger.debug(f"PUBLISHED detectado: ano na linha {check_row}")
                        return True
                        
        # SINAL 3: Verifica se há cabeçalhos típicos em colunas baixas
        for check_col in [1, 2]:
            if check_col <= ws.max_column:
                cell = ws.cell(row=data_row, column=check_col)
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower()
                    if any(pattern in cell_text for pattern in ['categoria', 'estabelecimentos', 'total']):
                        self.logger.debug(f"PUBLISHED detectado: padrão '{cell.value}' na coluna {check_col}")
                        return True
        
        # SINAL 4: Dados em linhas altas (30+) = provavelmente Published
        if data_row >= 30:
            self.logger.debug(f"PUBLISHED detectado: dados na linha {data_row} (muito alta)")
            return True
            
        # Default: assume recreated
        self.logger.debug(f"RECREATED detectado: linha {data_row} sem padrões published")
        return False

    def is_year(self, value) -> bool:
        """
        NOVO: Verifica se um valor é um ano (1900-2030).
        
        Args:
            value: Valor a verificar
            
        Returns:
            True se for um ano
        """
        try:
            if isinstance(value, str):
                num_value = float(value)
            elif isinstance(value, (int, float)):
                num_value = float(value)
            else:
                return False
                
            return 1900 <= num_value <= 2030
        except:
            return False

    def is_data_cell(self, cell) -> bool:
        """
        MELHORADO: Verifica se uma célula contém dados estatísticos legítimos.
        
        Aceita: Valores numéricos que são claramente dados estatísticos
        Rejeita: Texto, anos isolados, células vazias, cabeçalhos óbvios
        
        CRÍTICO: Este método deve identificar os valores que são DADOS,
        não cabeçalhos que vão ser usados para localizar esses dados.
        
        Args:
            cell: Célula a verificar
            
        Returns:
            True se for célula de dados estatísticos
        """
        if not cell.value:
            return False
            
        # Verifica se é numérico
        normalized_value = self.normalize_value(cell.value)
        if normalized_value is None:
            return False
            
        # CRÍTICO: Rejeita anos (são cabeçalhos de coluna, não dados)
        if self.is_year(cell.value):
            self.logger.debug(f"Rejeitado ano {cell.value} como dados - é cabeçalho de coluna")
            return False
            
        # CRÍTICO: Rejeita se for identificado como cabeçalho óbvio
        if self._is_likely_header(cell):
            self.logger.debug(f"Rejeitado {cell.value} como dados - é cabeçalho óbvio")
            return False
            
        # CORRIGIDO: Aceita valores estatísticos legítimos (incluindo decimais pequenos)
        # Rejeita apenas valores muito pequenos que são códigos óbvios (≤1)
        if normalized_value <= 1:
            self.logger.debug(f"Rejeitado código/indicador {normalized_value} - muito pequeno")
            return False
            
        # ACEITA todos os valores >1, incluindo decimais como 7.1, 6.8, etc.
        self.logger.debug(f"Aceite valor {normalized_value} como dados estatísticos")
        return True

    def extract_simple_data_points(self, file_path: str, sheet_name: str, file_type: str) -> List[Dict[str, Any]]:
        """
        MELHORADO: Extrai pontos de dados usando busca direcional simples.
        
        REJEITA: Dashes, símbolos, placeholders, valores estatísticos como cabeçalhos
        ACEITA: Apenas cabeçalhos de texto legítimo ou anos válidos
        APLICA: Validação rigorosa em cada coordenada extraída
        
        Para cada célula de dados:
        1. Anda ESQUERDA para encontrar row header REAL
        2. Anda CIMA para encontrar column header REAL  
        3. VALIDA se as coordenadas são legítimas
        4. Só adiciona se passou na validação
        
        Args:
            file_path: Caminho do arquivo Excel
            sheet_name: Nome da planilha
            file_type: 'published' ou 'recreated'
            
        Returns:
            Lista de pontos de dados com coordenadas validadas
        """
        data_points = []
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            
            total_extracted = 0
            total_rejected = 0
            
            self.logger.info(f"🔍 Iniciando extração simples para {file_type} - {sheet_name}")
            self.logger.info(f"   Dimensões da planilha: {ws.max_row} linhas x {ws.max_column} colunas")
            
            # Processa cada célula da planilha
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Verifica se é uma célula de dados válida
                    if self.is_data_cell(cell):
                        total_extracted += 1
                        value = self.get_displayed_value(cell)
                        
                        if value is None:
                            continue
                            
                        # BUSCA DIRECIONAL: Encontra cabeçalhos
                        row_header = self.find_row_header(ws, row, col)
                        column_header = self.find_column_header(ws, row, col)
                        
                        # CRÍTICO: Valida as coordenadas antes de adicionar
                        if not self.validate_coordinate_pair(row_header, column_header, value):
                            total_rejected += 1
                            self.logger.debug(f"REJEITADO ponto inválido: ({row_header}, {column_header}) = {value}")
                            continue
                        
                        # APLICA equivalência semântica
                        row_header_normalized = self.apply_semantic_equivalence(row_header)
                        column_header_normalized = self.apply_semantic_equivalence(column_header)
                        
                        data_point = {
                            'row': row_header_normalized,
                            'column': column_header_normalized,
                            'value': value,
                            'position': (row, col),
                            'original_row': row_header,
                            'original_column': column_header
                        }
                        
                        data_points.append(data_point)
                        
                        # Log para debug a cada 100 pontos válidos
                        if len(data_points) % 100 == 0:
                            self.logger.debug(f"   Progresso: {len(data_points)} pontos válidos extraídos...")
            
            # RESUMO da extração
            self.logger.info(f"✅ Extração simples concluída para {file_type} - {sheet_name}:")
            self.logger.info(f"   Células de dados encontradas: {total_extracted}")
            self.logger.info(f"   Pontos rejeitados na validação: {total_rejected}")
            self.logger.info(f"   Pontos válidos extraídos: {len(data_points)}")
            
            if total_rejected > 0:
                rejection_rate = (total_rejected / total_extracted) * 100 if total_extracted > 0 else 0
                self.logger.warning(f"   ⚠️ Taxa de rejeição: {rejection_rate:.1f}% ({total_rejected}/{total_extracted})")
                
            if len(data_points) == 0:
                self.logger.error(f"❌ PROBLEMA: Nenhum ponto de dados válido extraído para {file_type}!")
                self.logger.error(f"   Isto indica problemas na detecção de cabeçalhos.")
                
            # Mostra exemplos dos primeiros pontos para debug
            if data_points:
                self.logger.debug(f"   Exemplos de coordenadas extraídas:")
                for i, point in enumerate(data_points[:5]):
                    self.logger.debug(f"     {i+1}. ({point['row']}, {point['column']}) = {point['value']}")
                if len(data_points) > 5:
                    self.logger.debug(f"     ... mais {len(data_points) - 5} pontos")
            
            wb.close()
            return data_points
            
        except Exception as e:
            self.logger.error(f"Erro na extração simples de {file_type}: {e}", exc_info=True)
            return []

    def apply_semantic_equivalence(self, header_text: str) -> str:
        """
        NOVO: Aplica mapeamentos semânticos simples durante extração de cabeçalhos.
        
        Args:
            header_text: Texto do cabeçalho
            
        Returns:
            Texto do cabeçalho com equivalências aplicadas
        """
        if not header_text:
            return header_text
            
        equivalence_map = {
            "****": "4", "***": "3", "**": "2", "*": "1",
            "Total": "(Em branco)",
            "(Em branco)": "Total", 
            "(em branco)": "Total",
            "total": "Total",
            "TOTAL": "Total",
            "Estabelecimentos hoteleiros": "Hotelaria",
            "Hotéis": "Hotel",
            "Pensões": "Pensão", 
            "Motéis": "Motel",
            "Hotelaria": "Hotelaria",  # Normalização
            "Hotel": "Hotel",
            "Pensão": "Pensão",
            "Motel": "Motel"
        }
        
        clean_text = str(header_text).strip()
        result = equivalence_map.get(clean_text, clean_text)
        
        if result != clean_text:
            self.logger.debug(f"Equivalência semântica aplicada: '{clean_text}' -> '{result}'")
            
        return result

    def compare_simple_data(self, published_points: List[Dict[str, Any]], 
                           recreated_points: List[Dict[str, Any]], 
                           sheet_name: str) -> Dict[str, Any]:
        """
        CORRIGIDO: Compara APENAS coordenadas lógicas (row, column) - NÃO posições físicas!
        
        Posições são diferentes entre ficheiros - devem ser usadas apenas 
        para highlighting, NÃO para lógica de comparação.
        
        Args:
            published_points: Pontos de dados do ficheiro publicado
            recreated_points: Pontos de dados do ficheiro recriado
            sheet_name: Nome da folha
            
        Returns:
            Dicionário com resultados da comparação
        """
        self.logger.info(f"[COMPARAÇÃO] Folha '{sheet_name}': {len(published_points)} publicados vs {len(recreated_points)} recriados")
        
        # DEBUG: Log coordenadas de exemplo de cada ficheiro
        self.logger.info(f"[DEBUG] Coordenadas de exemplo PUBLICADAS:")
        for i, point in enumerate(published_points[:5]):
            key = (point['row'], point['column'])
            self.logger.info(f"  {i}: {key} = {point['value']} (posição: {point['position']})")
        
        self.logger.info(f"[DEBUG] Coordenadas de exemplo RECRIADAS:")
        for i, point in enumerate(recreated_points[:5]):
            key = (point['row'], point['column'])
            self.logger.info(f"  {i}: {key} = {point['value']} (posição: {point['position']})")
        
        # CRÍTICO: Constrói mapas usando APENAS (row, column) com normalização semântica!
        published_map = {}
        for point in published_points:
            key = self.normalize_coordinate_key(point['row'], point['column'])  # COM NORMALIZAÇÃO!
            published_map[key] = point
            
        recreated_map = {}
        for point in recreated_points:
            key = self.normalize_coordinate_key(point['row'], point['column'])  # COM NORMALIZAÇÃO!
            recreated_map[key] = point
        
        # DEBUG: Encontra chaves comuns
        common_keys = set(published_map.keys()).intersection(set(recreated_map.keys()))
        self.logger.info(f"[DEBUG] Chaves de coordenadas comuns encontradas: {len(common_keys)}")
        
        if len(common_keys) > 0:
            self.logger.info(f"[DEBUG] Primeiras 3 chaves comuns: {list(common_keys)[:3]}")
        else:
            self.logger.error(f"[DEBUG] NENHUMA CHAVE COMUM! Isto explica 0% de precisão")
            self.logger.error(f"[DEBUG] Chaves publicadas (amostra): {list(published_map.keys())[:3]}")
            self.logger.error(f"[DEBUG] Chaves recriadas (amostra): {list(recreated_map.keys())[:3]}")
        
        results = {
            'correct_matches': [],
            'value_differences': [],
            'missing_in_published': [],
            'missing_in_recreated': [],
            'sheet_name': sheet_name,
            'published_data_points': len(published_points),
            'recreated_data_points': len(recreated_points),
            'published_count': len(published_points),  # Mantém compatibilidade
            'recreated_count': len(recreated_points)   # Mantém compatibilidade
        }
        
        # Tolerância numérica para comparação
        tolerance = self.numeric_tolerance
        
        # Compara recriados contra publicados usando coordenadas lógicas
        for key, recreated_point in recreated_map.items():
            if key in published_map:  # Mesma coordenada lógica (row, column)
                published_point = published_map[key]
                
                # Compara valores (posições são irrelevantes)
                value_diff = abs(recreated_point['value'] - published_point['value'])
                
                if value_diff <= tolerance:
                    results['correct_matches'].append({
                        'coordinates': key,  # (row, column) apenas
                        'published_value': published_point['value'],
                        'recreated_value': recreated_point['value'],
                        'position': recreated_point['position'],  # Para highlighting apenas
                        'difference': value_diff
                    })
                else:
                    results['value_differences'].append({
                        'coordinates': key,  # (row, column) apenas
                        'published_value': published_point['value'],
                        'recreated_value': recreated_point['value'],
                        'position': recreated_point['position'],  # Para highlighting apenas
                        'difference': recreated_point['value'] - published_point['value']
                    })
            else:
                results['missing_in_published'].append({
                    'coordinates': key,  # (row, column) apenas
                    'recreated_value': recreated_point['value'],
                    'position': recreated_point['position']  # Para highlighting apenas
                })
        
        # Encontra valores apenas no publicado
        for key, published_point in published_map.items():
            if key not in recreated_map:
                results['missing_in_recreated'].append({
                    'coordinates': key,  # (row, column) apenas
                    'published_value': published_point['value']
                })
        
        # Estatísticas da comparação
        total_comparisons = len(recreated_points)
        correct_matches = len(results['correct_matches'])
        value_differences = len(results['value_differences'])
        missing_in_published = len(results['missing_in_published'])
        missing_in_recreated = len(results['missing_in_recreated'])
        
        accuracy = (correct_matches / total_comparisons * 100) if total_comparisons > 0 else 0
        
        self.logger.info(f"[COMPARAÇÃO] Folha '{sheet_name}' - Resultados:")
        self.logger.info(f"  ✅ Correspondências correctas: {correct_matches}")
        self.logger.info(f"  ❌ Diferenças de valor: {value_differences}")
        self.logger.info(f"  ❓ Ausentes no publicado: {missing_in_published}")
        self.logger.info(f"  ❓ Ausentes no recriado: {missing_in_recreated}")
        self.logger.info(f"  📊 Precisão: {accuracy:.2f}%")
        
        results['accuracy'] = accuracy
        # CORRIGIDO: Discrepâncias = apenas problemas visíveis no ficheiro recriado
        results['total_discrepancies'] = value_differences + missing_in_published
        results['discrepancy_count'] = value_differences + missing_in_published  # SEM missing_in_recreated
        
        return results

    def apply_enhanced_highlighting(self, target_ws, sheet_name: str, comparison_results: Dict[str, Any]) -> Tuple[int, int]:
        """
        CORRIGIDO: Aplica destaque aprimorado com sistema de duas cores.
        
        Verde: Correspondências correctas
        Amarelo: Discrepâncias ou dados ausentes
        
        Remove referências a 'original_cell' que não existe na nova estrutura.
        
        Args:
            target_ws: Worksheet de destino
            sheet_name: Nome da folha
            comparison_results: Resultados da comparação
            
        Returns:
            Tupla com (células destacadas, comentários adicionados)
        """
        highlighted_cells = 0
        comments_added = 0
        
        try:
            from openpyxl.styles import PatternFill
            from openpyxl.comments import Comment
            
            # Cores para destaque
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Amarelo claro
            
            # Destaca correspondências correctas (VERDE)
            matches = comparison_results.get('correct_matches', [])
            for match in matches:
                try:
                    # USA POSIÇÃO em vez de original_cell
                    if 'position' in match:
                        row, col = match['position']
                        cell = target_ws.cell(row=row, column=col)
                        cell.fill = green_fill
                        
                        # Adiciona comentário de sucesso
                        comment_text = self.create_success_comment(match)
                        cell.comment = Comment(comment_text, "Sistema ETL")
                        
                        highlighted_cells += 1
                        comments_added += 1
                        
                except Exception as e:
                    self.logger.warning(f"Erro ao destacar correspondência: {e}")
            
            # Destaca discrepâncias (AMARELO)
            differences = comparison_results.get('value_differences', [])
            for diff in differences:
                try:
                    # USA POSIÇÃO em vez de original_cell
                    if 'position' in diff:
                        row, col = diff['position']
                        cell = target_ws.cell(row=row, column=col)
                        cell.fill = yellow_fill
                        
                        # Adiciona comentário de erro
                        comment_text = self.create_error_comment(diff)
                        cell.comment = Comment(comment_text, "Sistema ETL")
                        
                        highlighted_cells += 1
                        comments_added += 1
                        
                except Exception as e:
                    self.logger.warning(f"Erro ao destacar discrepância: {e}")
            
            # Destaca dados ausentes (AMARELO)
            missing_published = comparison_results.get('missing_in_published', [])
            missing_recreated = comparison_results.get('missing_in_recreated', [])
            
            for missing in missing_published + missing_recreated:
                try:
                    # USA POSIÇÃO em vez de original_cell
                    if 'position' in missing:
                        row, col = missing['position']
                        cell = target_ws.cell(row=row, column=col)
                        cell.fill = yellow_fill
                        
                        # Adiciona comentário de ausência
                        comment_text = self.create_missing_comment(missing)
                        cell.comment = Comment(comment_text, "Sistema ETL")
                        
                        highlighted_cells += 1
                        comments_added += 1
                        
                except Exception as e:
                    self.logger.warning(f"Erro ao destacar dado ausente: {e}")
            
            self.logger.info(f"[{sheet_name}] Destaque aplicado: {highlighted_cells} células, {comments_added} comentários")
            return highlighted_cells, comments_added
            
        except Exception as e:
            self.logger.error(f"Erro durante aplicação de destaque: {e}")
            return 0, 0

    def create_success_comment(self, match: Dict[str, Any]) -> str:
        """NOVO: Cria comentário simplificado para correspondências correctas"""
        row_name, col_name = match['coordinates']
        
        comment = "=== COMPARAÇÃO REALIZADA ===\n"
        comment += f"Posição: {col_name} x {row_name}"
        
        return comment

    def create_error_comment(self, diff: Dict[str, Any]) -> str:
        """NOVO: Cria comentário simplificado para diferenças de valor"""
        row_name, col_name = diff['coordinates']
        
        comment = "=== DISCREPÂNCIA DETECTADA ===\n"
        comment += f"Posição: {col_name} x {row_name}"
        
        return comment

    def create_missing_comment(self, missing: Dict[str, Any]) -> str:
        """NOVO: Cria comentário simplificado para valores não encontrados"""
        row_name, col_name = missing['coordinates']
        
        comment = "=== VALOR EM FALTA ===\n"
        comment += f"Posição: {col_name} x {row_name}"
        
        return comment

    def normalize_coordinate_key(self, row: str, column: str) -> Tuple[str, str]:
        """
        NOVO: Aplica equivalência semântica ao criar chaves de comparação.
        
        Args:
            row: Cabeçalho da linha
            column: Cabeçalho da coluna
            
        Returns:
            Tupla com coordenadas normalizadas
        """
        normalized_row = self.apply_semantic_equivalence(row)
        normalized_column = self.apply_semantic_equivalence(column)
        return (normalized_row, normalized_column)

    def validate_coordinate_pair(self, row_header: str, column_header: str, value: float) -> bool:
        """
        MELHORADO: Valida se um par de coordenadas é legítimo ou suspeito.
        
        Detecta quando valores estatísticos, dashes, símbolos ou placeholders
        estão a ser incorretamente usados como cabeçalhos.
        
        Args:
            row_header: Cabeçalho da linha extraído
            column_header: Cabeçalho da coluna extraído
            value: Valor dos dados
            
        Returns:
            True se as coordenadas parecem legítimas
        """
        
        # CRÍTICO: Lista de cabeçalhos inválidos
        invalid_headers = {'-', '.', '•', '*', '**', '***', '****', 
                          '_', '()', '( )', ':', ';', ',', '|', 
                          '/', '\\', '+', '=', '#', '@', '!', '?'}
        
        # CRÍTICO: Rejeita dashes e símbolos como cabeçalhos
        if row_header in invalid_headers:
            self.logger.error(f"❌ COORDENADA INVÁLIDA: Row header '{row_header}' é símbolo/placeholder!")
            self.logger.error(f"   Valor associado: {value}, Column: {column_header}")
            return False
            
        if column_header in invalid_headers:
            self.logger.error(f"❌ COORDENADA INVÁLIDA: Column header '{column_header}' é símbolo/placeholder!")
            self.logger.error(f"   Valor associado: {value}, Row: {row_header}")
            return False
        
        # CRÍTICO: Detecta se o row_header é um valor estatístico
        try:
            row_as_number = float(row_header.replace(',', '.'))
            if row_as_number > 1000:  # Valor estatístico usado como row header
                self.logger.error(f"❌ COORDENADA SUSPEITA: Row header '{row_header}' parece ser valor estatístico!")
                self.logger.error(f"   Valor associado: {value}, Column: {column_header}")
                return False
        except ValueError:
            pass  # É texto - OK
            
        # CRÍTICO: Detecta se o column_header é um valor estatístico (exceto anos)
        try:
            col_as_number = float(column_header.replace(',', '.'))
            if col_as_number > 2030:  # Valor estatístico usado como column header
                self.logger.error(f"❌ COORDENADA SUSPEITA: Column header '{column_header}' parece ser valor estatístico!")
                self.logger.error(f"   Valor associado: {value}, Row: {row_header}")
                return False
        except ValueError:
            pass  # É texto - OK
            
        # CRÍTICO: Rejeita cabeçalhos muito curtos (menos de 2 caracteres)
        if len(row_header.strip()) < 2 and not row_header.startswith('Row_'):
            self.logger.error(f"❌ COORDENADA INVÁLIDA: Row header '{row_header}' muito curto!")
            return False
            
        if len(column_header.strip()) < 2 and not column_header.startswith('Col_'):
            self.logger.error(f"❌ COORDENADA INVÁLIDA: Column header '{column_header}' muito curto!")
            return False
            
        # AVISO: Detecta uso de fallbacks
        # AVISO: Detecta uso de fallbacks
        if row_header.startswith('Row_') or column_header.startswith('Col_'):
            self.logger.warning(f"⚠️ FALLBACK usado: ({row_header}, {column_header}) para valor {value}")
            return True  # Aceita, mas avisa
            
        # Coordenadas parecem legítimas
        self.logger.debug(f"✅ Coordenadas válidas: ({row_header}, {column_header}) = {value}")
        return True


def run_interactive_comparison(logger: logging.Logger):
    """
    Executa o processo interativo de comparação de dados.
    
    Args:
        logger: Logger configurado
    """
    from colorama import Fore, Style
    
    try:
        # Cria instância do comparador
        comparator = DataComparator(logger)
        
        print(f"\n{Fore.GREEN}[Comparação Inteligente de Dados Excel]{Style.RESET_ALL}")
        print("Esta funcionalidade compara ficheiros Excel com estruturas de tabela cruzada,")
        print("aplicando normalização avançada e correspondência difusa de dimensões.\n")
        
        # Seleção interativa de ficheiros
        published_file, recreated_file = comparator.select_files_interactively()
        
        if not published_file or not recreated_file:
            print(f"{Fore.YELLOW}Operação cancelada.{Style.RESET_ALL}")
            return
        
        # Seleção de folhas
        selected_sheets = comparator.select_sheets_interactively(published_file, recreated_file)
        
        if not selected_sheets:
            print(f"{Fore.YELLOW}Nenhuma folha selecionada para comparação.{Style.RESET_ALL}")
            return
        
        # Confirmação
        print(f"\n{Fore.CYAN}Configuração da Comparação:{Style.RESET_ALL}")
        print(f"Ficheiro publicado: {published_file}")
        print(f"Ficheiro recriado: {recreated_file}")
        print(f"Folhas a comparar: {', '.join(selected_sheets)}")
        
        confirm = input(f"\n{Fore.GREEN}Continuar com a comparação? (s/N):{Style.RESET_ALL} ").strip().lower()
        if confirm not in ['s', 'sim', 'y', 'yes']:
            print(f"{Fore.YELLOW}Operação cancelada.{Style.RESET_ALL}")
            return
        
        # Executa comparação
        print(f"\n{Fore.CYAN}Iniciando comparação...{Style.RESET_ALL}")
        results = comparator.compare_files(published_file, recreated_file, selected_sheets)
        
        # Apresenta resumo
        print(f"\n{Fore.GREEN}Comparação concluída!{Style.RESET_ALL}")
        print(f"\n{Fore.CYAN}Resumo dos Resultados:{Style.RESET_ALL}")
        
        summary = results['summary']
        print(f"• Total de pontos publicados: {summary['total_published_points']}")
        print(f"• Total de pontos recriados: {summary['total_recreated_points']}")
        print(f"• Total de discrepâncias: {summary['total_discrepancies']}")
        print(f"• Precisão geral: {summary['accuracy_percentage']:.2f}%")
        
        # Resumo por folha
        if results['sheet_results']:
            print(f"\n{Fore.CYAN}Resumo por Folha:{Style.RESET_ALL}")
            for sheet_name, sheet_results in results['sheet_results'].items():
                if 'error' not in sheet_results:
                    # CORRIGIDO: Usa EXACTAMENTE os mesmos dados que vão para o Excel
                    correct_matches = len(sheet_results.get('correct_matches', []))
                    value_differences = len(sheet_results.get('value_differences', []))
                    missing_in_published = len(sheet_results.get('missing_in_published', []))
                    missing_in_recreated = len(sheet_results.get('missing_in_recreated', []))
                    
                    # CORRIGIDO: Pontos totais de dados (published + recreated)
                    published_points = sheet_results.get('published_data_points', 0)
                    recreated_points = sheet_results.get('recreated_data_points', 0)
                    total_data_points = published_points + recreated_points
                    
                    # CORRIGIDO: Total de comparações feitas (apenas recreated, pois é isso que comparamos)
                    total_comparisons = recreated_points
                    # CORRIGIDO: Discrepâncias = apenas problemas visíveis no ficheiro recriado
                    discrepancies = value_differences + missing_in_published  # SEM missing_in_recreated
                    
                    # CORRIGIDO: Precisão baseada nas correspondências sobre comparações
                    if total_comparisons > 0:
                        accuracy = (correct_matches / total_comparisons) * 100
                    else:
                        accuracy = 0.0
                    
                    status_color = Fore.GREEN if accuracy > 95 else Fore.YELLOW if accuracy > 80 else Fore.RED
                    print(f"• {sheet_name}: {status_color}{accuracy:.2f}%{Style.RESET_ALL} "
                          f"({discrepancies} discrepâncias, {published_points} pub, {recreated_points} rec)")
                else:
                    print(f"• {sheet_name}: {Fore.RED}Erro{Style.RESET_ALL} - {sheet_results['error']}")
        
        # Gera relatório
        print(f"\n{Fore.CYAN}Gerando relatório detalhado...{Style.RESET_ALL}")
        report_file = comparator.generate_comparison_report(results)
        
        print(f"\n{Fore.GREEN}✅ Relatório gerado com sucesso!{Style.RESET_ALL}")
        print(f"Ficheiro: {report_file}")
        
        # Mostra próximos passos
        if summary['total_discrepancies'] > 0:
            print(f"\n{Fore.YELLOW}Próximos Passos:{Style.RESET_ALL}")
            print("1. Revise o relatório Excel gerado para detalhes das discrepâncias")
            print("2. Verifique se as diferenças são aceitáveis ou requerem correção")
            print("3. Ajuste os dados recriados conforme necessário")
        else:
            print(f"\n{Fore.GREEN}✅ Perfeito!{Style.RESET_ALL} Nenhuma discrepância encontrada.")
            print("Os dados recriados correspondem exatamente aos publicados.")
        
    except Exception as e:
        logger.error(f"Erro durante comparação: {e}", exc_info=True)
        print(f"\n{Fore.RED}Erro crítico durante a comparação:{Style.RESET_ALL} {str(e)}")
        print("Consulte os logs para detalhes técnicos.")