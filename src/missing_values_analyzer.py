#!/usr/bin/env python3
"""
Analisador de Valores em Falta - Sistema ETL
===========================================

Este módulo é especializado na detecção de valores em falta no dataset recriado
quando comparado com o dataset publicado. Focado exclusivamente em identificar
dados que existem no ficheiro publicado mas estão ausentes no recriado.

Características principais:
- Detecção robusta de coordenadas em falta
- Relatório Excel organizado e claro
- Interface de console consistente com validador
- Análise folha por folha detalhada
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Tuple, Optional
from colorama import Fore, Style

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️ AVISO: openpyxl não está instalado. Funcionalidade Excel limitada.")

from src.data_comparator import DataComparator


class MissingValuesAnalyzer:
    """
    Analisador especializado em valores em falta.
    
    Detecta coordenadas que existem no ficheiro publicado mas estão
    ausentes no ficheiro recriado, fornecendo relatórios detalhados.
    """
    
    def __init__(self, logger: logging.Logger):
        """
        Inicializa o analisador de valores em falta.
        
        Args:
            logger: Logger configurado
        """
        self.logger = logger
        self.comparator = DataComparator(logger)
        
    def get_available_files(self) -> Tuple[List[str], Dict[str, List[str]]]:
        """
        Herda funcionalidade do DataComparator para manter consistência.
        
        Returns:
            Tupla com (todos os ficheiros Excel, ficheiros por diretório)
        """
        return self.comparator.get_available_files()
    
    def select_files_interactively(self) -> Tuple[Optional[str], Optional[str]]:
        """
        Interface interativa para seleção de ficheiros.
        
        Returns:
            Tupla com (ficheiro_publicado, ficheiro_recriado) ou (None, None)
        """
        display_header()
        print(f"{Fore.GREEN}[Seleção de Ficheiros para Análise de Valores em Falta]{Style.RESET_ALL}")
        print("Selecione os ficheiros para comparação:\n")
        
        # Usa a mesma lógica do comparador mas com texto específico
        return self.comparator.select_files_interactively()
    
    def select_sheets_interactively(self, published_file: str, recreated_file: str) -> List[str]:
        """
        Interface para seleção de folhas a analisar.
        
        Args:
            published_file: Caminho do ficheiro publicado
            recreated_file: Caminho do ficheiro recriado
            
        Returns:
            Lista de nomes das folhas selecionadas
        """
        return self.comparator.select_sheets_interactively(published_file, recreated_file)
    
    def analyze_missing_values(self, published_file: str, recreated_file: str, 
                              sheet_names: List[str]) -> Dict[str, Any]:
        """
        Analisa valores em falta entre ficheiros publicado e recriado.
        
        Args:
            published_file: Caminho do ficheiro publicado
            recreated_file: Caminho do ficheiro recriado
            sheet_names: Lista de folhas a analisar
            
        Returns:
            Dicionário com resultados da análise
        """
        self.logger.info(f"🔍 Iniciando análise de valores em falta")
        self.logger.info(f"Ficheiro publicado: {published_file}")
        self.logger.info(f"Ficheiro recriado: {recreated_file}")
        self.logger.info(f"Folhas a analisar: {sheet_names}")
        
        results = {
            'published_file': published_file,
            'recreated_file': recreated_file,
            'sheet_results': {},
            'summary': {}
        }
        
        total_missing_values = 0
        total_published_points = 0
        
        # Analisa cada folha
        for sheet_name in sheet_names:
            self.logger.info(f"📊 Analisando folha: {sheet_name}")
            
            try:
                sheet_result = self.analyze_sheet_missing_values(
                    published_file, recreated_file, sheet_name
                )
                results['sheet_results'][sheet_name] = sheet_result
                
                if 'error' not in sheet_result:
                    total_missing_values += len(sheet_result.get('missing_values', []))
                    total_published_points += sheet_result.get('published_count', 0)
                    
            except Exception as e:
                self.logger.error(f"Erro ao analisar folha {sheet_name}: {e}")
                results['sheet_results'][sheet_name] = {'error': str(e)}
        
        # Calcula estatísticas gerais
        missing_percentage = (total_missing_values / total_published_points * 100) if total_published_points > 0 else 0
        
        results['summary'] = {
            'total_published_points': total_published_points,
            'total_missing_values': total_missing_values,
            'missing_percentage': missing_percentage,
            'analysis_timestamp': datetime.now().isoformat()
        }
        
        self.logger.info(f"✅ Análise concluída:")
        self.logger.info(f"  📊 Total de pontos publicados: {total_published_points}")
        self.logger.info(f"  ❌ Total de valores em falta: {total_missing_values}")
        self.logger.info(f"  📈 Percentagem em falta: {missing_percentage:.2f}%")
        
        return results
    
    def analyze_sheet_missing_values(self, published_file: str, recreated_file: str, 
                                    sheet_name: str) -> Dict[str, Any]:
        """
        Analisa valores em falta numa folha específica.
        
        Args:
            published_file: Caminho do ficheiro publicado
            recreated_file: Caminho do ficheiro recriado
            sheet_name: Nome da folha
            
        Returns:
            Dicionário com resultados da folha
        """
        try:
            # Extrai pontos de dados de ambos os ficheiros
            published_points = self.comparator.extract_simple_data_points(
                published_file, sheet_name, 'published'
            )
            recreated_points = self.comparator.extract_simple_data_points(
                recreated_file, sheet_name, 'recreated'
            )
            
            self.logger.info(f"[FALTA] Folha '{sheet_name}': {len(published_points)} publicados, {len(recreated_points)} recriados")
            
            # Cria mapas de coordenadas
            published_map = {}
            for point in published_points:
                key = self.comparator.normalize_coordinate_key(point['row'], point['column'])
                published_map[key] = point
                
            recreated_map = {}
            for point in recreated_points:
                key = self.comparator.normalize_coordinate_key(point['row'], point['column'])
                recreated_map[key] = point
            
            # Encontra valores em falta (existem no publicado mas não no recriado)
            missing_values = []
            for key, published_point in published_map.items():
                if key not in recreated_map:
                    missing_values.append({
                        'coordinates': key,
                        'published_value': published_point['value'],
                        'row_header': key[0],
                        'column_header': key[1]
                    })
            
            # Calcula estatísticas
            missing_count = len(missing_values)
            published_count = len(published_points)
            missing_percentage = (missing_count / published_count * 100) if published_count > 0 else 0
            
            self.logger.info(f"[FALTA] Folha '{sheet_name}' - Resultados:")
            self.logger.info(f"  📊 Pontos publicados: {published_count}")
            self.logger.info(f"  ❌ Valores em falta: {missing_count}")
            self.logger.info(f"  📈 Percentagem em falta: {missing_percentage:.2f}%")
            
            return {
                'sheet_name': sheet_name,
                'missing_values': missing_values,
                'published_count': published_count,
                'recreated_count': len(recreated_points),
                'missing_count': missing_count,
                'missing_percentage': missing_percentage
            }
            
        except Exception as e:
            self.logger.error(f"Erro ao analisar folha {sheet_name}: {e}")
            return {'error': str(e)}
    
    def generate_missing_values_report(self, results: Dict[str, Any], 
                                     output_dir: str = "result/missing_analysis") -> str:
        """
        Gera relatório Excel dos valores em falta.
        
        Args:
            results: Resultados da análise
            output_dir: Diretório de saída
            
        Returns:
            Caminho do ficheiro gerado
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = f"analise_valores_em_falta_{timestamp}.xlsx"
            report_path = os.path.join(output_dir, report_filename)
            
            self.logger.info(f"📝 Gerando relatório: {report_path}")
            
            # Cria workbook
            wb = Workbook()
            
            # Remove folha padrão
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Cria folha de resumo geral
            summary_ws = wb.create_sheet("Resumo_Geral")
            self._create_missing_summary_sheet(summary_ws, results)
            
            # Cria folhas detalhadas para cada folha analisada
            for sheet_name, sheet_results in results['sheet_results'].items():
                if 'error' not in sheet_results and sheet_results.get('missing_count', 0) > 0:
                    detail_ws = wb.create_sheet(f"Falta_{sheet_name}")
                    self._create_missing_detail_sheet(detail_ws, sheet_results)
            
            # Cria folha de informações técnicas
            tech_ws = wb.create_sheet("Info_Tecnica")
            self._create_missing_technical_sheet(tech_ws, results)
            
            # Guarda ficheiro
            wb.save(report_path)
            self.logger.info(f"✅ Relatório guardado: {report_path}")
            
            return report_path
            
        except Exception as e:
            self.logger.error(f"Erro ao gerar relatório: {e}")
            raise
    
    def _create_missing_summary_sheet(self, ws, results: Dict[str, Any]):
        """Cria folha de resumo geral dos valores em falta."""
        # Cabeçalho
        ws['A1'] = "RELATÓRIO DE ANÁLISE DE VALORES EM FALTA"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        ws[f'A{row}'] = "Ficheiro Publicado:"
        ws[f'B{row}'] = results['published_file']
        row += 1
        
        ws[f'A{row}'] = "Ficheiro Recriado:"
        ws[f'B{row}'] = results['recreated_file']
        row += 1
        
        ws[f'A{row}'] = "Data da Análise:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 2
        
        # Estatísticas gerais
        ws[f'A{row}'] = "ESTATÍSTICAS GERAIS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        summary = results['summary']
        ws[f'A{row}'] = "Total de pontos publicados:"
        ws[f'B{row}'] = summary['total_published_points']
        row += 1
        
        ws[f'A{row}'] = "Total de valores em falta:"
        ws[f'B{row}'] = summary['total_missing_values']
        row += 1
        
        ws[f'A{row}'] = "Percentagem em falta:"
        ws[f'B{row}'] = f"{summary['missing_percentage']:.2f}%"
        row += 2
        
        # Resumo por folha
        ws[f'A{row}'] = "RESUMO POR FOLHA"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Cabeçalhos da tabela
        headers = ["Folha", "Pontos Publicados", "Valores em Falta", "Percentagem (%)", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        
        # Dados por folha
        for sheet_name, sheet_results in results['sheet_results'].items():
            if 'error' not in sheet_results:
                published_count = sheet_results.get('published_count', 0)
                missing_count = sheet_results.get('missing_count', 0)
                missing_percentage = sheet_results.get('missing_percentage', 0.0)
                
                status = "✅ Completo" if missing_count == 0 else f"❌ {missing_count} em falta"
                
                ws[f'A{row}'] = sheet_name
                ws[f'B{row}'] = published_count
                ws[f'C{row}'] = missing_count
                ws[f'D{row}'] = f"{missing_percentage:.2f}%"
                ws[f'E{row}'] = status
                
                # Destaca folhas com problemas
                if missing_count > 0:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )
                
                row += 1
        
        # Ajusta largura das colunas
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_missing_detail_sheet(self, ws, sheet_results: Dict[str, Any]):
        """Cria folha detalhada com todos os valores em falta."""
        sheet_name = sheet_results['sheet_name']
        missing_values = sheet_results.get('missing_values', [])
        
        ws['A1'] = f"VALORES EM FALTA - {sheet_name}"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = f"Total de valores em falta: {len(missing_values)}"
        ws['A3'].font = Font(bold=True)
        
        if not missing_values:
            ws['A5'] = "✅ Nenhum valor em falta nesta folha."
            return
        
        # Cabeçalhos da tabela
        headers = ["#", "Coordenadas", "Linha", "Coluna", "Valor Publicado"]
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados dos valores em falta
        for i, missing in enumerate(missing_values, 1):
            row += 1
            coordinates = f"({missing['row_header']}, {missing['column_header']})"
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=coordinates)
            ws.cell(row=row, column=3, value=missing['row_header'])
            ws.cell(row=row, column=4, value=missing['column_header'])
            ws.cell(row=row, column=5, value=missing['published_value'])
            
            # Destaca linha
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        
        # Ajusta largura das colunas
        column_widths = [8, 25, 15, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_missing_technical_sheet(self, ws, results: Dict[str, Any]):
        """Cria folha com informações técnicas da análise."""
        ws['A1'] = "DETALHES TÉCNICOS DA ANÁLISE"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Informações gerais
        ws[f'A{row}'] = "PARÂMETROS DA ANÁLISE"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Tipo de análise:"
        ws[f'B{row}'] = "Valores em Falta (Published → Recreated)"
        row += 1
        
        ws[f'A{row}'] = "Método de comparação:"
        ws[f'B{row}'] = "Coordenadas normalizadas com equivalência semântica"
        row += 2
        
        # Resumo por folha
        ws[f'A{row}'] = "ANÁLISE DETALHADA POR FOLHA"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for sheet_name, sheet_results in results['sheet_results'].items():
            ws[f'A{row}'] = f"Folha: {sheet_name}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if 'error' in sheet_results:
                ws[f'B{row}'] = f"❌ Erro: {sheet_results['error']}"
                row += 1
            else:
                ws[f'B{row}'] = f"📊 Pontos publicados: {sheet_results.get('published_count', 0)}"
                row += 1
                ws[f'B{row}'] = f"📊 Pontos recriados: {sheet_results.get('recreated_count', 0)}"
                row += 1
                ws[f'B{row}'] = f"❌ Valores em falta: {sheet_results.get('missing_count', 0)}"
                row += 1
                ws[f'B{row}'] = f"📈 Percentagem em falta: {sheet_results.get('missing_percentage', 0):.2f}%"
                row += 1
            row += 1
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50


def display_header():
    """Exibe cabeçalho específico para análise de valores em falta."""
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"\n{Fore.CYAN}{'='*80}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}  ANÁLISE DE VALORES EM FALTA - SISTEMA ETL{Style.RESET_ALL}")
    print(f"{Fore.CYAN}  Detecção de Dados Ausentes no Dataset Recriado{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'='*80}{Style.RESET_ALL}\n")


def run_missing_values_analysis(logger: logging.Logger):
    """
    Executa o processo interativo de análise de valores em falta.
    
    Args:
        logger: Logger configurado
    """
    try:
        # Cria instância do analisador
        analyzer = MissingValuesAnalyzer(logger)
        
        print(f"\n{Fore.GREEN}[Análise de Valores em Falta]{Style.RESET_ALL}")
        print("Esta funcionalidade identifica dados que existem no ficheiro publicado")
        print("mas estão ausentes no ficheiro recriado, fornecendo relatórios detalhados.\n")
        
        # Seleção interativa de ficheiros
        published_file, recreated_file = analyzer.select_files_interactively()
        
        if not published_file or not recreated_file:
            print(f"{Fore.YELLOW}Operação cancelada.{Style.RESET_ALL}")
            return
        
        # Seleção de folhas
        selected_sheets = analyzer.select_sheets_interactively(published_file, recreated_file)
        
        if not selected_sheets:
            print(f"{Fore.YELLOW}Nenhuma folha selecionada para análise.{Style.RESET_ALL}")
            return
        
        # Confirmação
        print(f"\n{Fore.CYAN}Configuração da Análise:{Style.RESET_ALL}")
        print(f"Ficheiro publicado: {published_file}")
        print(f"Ficheiro recriado: {recreated_file}")
        print(f"Folhas a analisar: {', '.join(selected_sheets)}")
        
        confirm = input(f"\n{Fore.GREEN}Continuar com a análise? (s/N):{Style.RESET_ALL} ").strip().lower()
        if confirm not in ['s', 'sim', 'y', 'yes']:
            print(f"{Fore.YELLOW}Operação cancelada.{Style.RESET_ALL}")
            return
        
        # Executa análise
        print(f"\n{Fore.CYAN}Iniciando análise de valores em falta...{Style.RESET_ALL}")
        results = analyzer.analyze_missing_values(published_file, recreated_file, selected_sheets)
        
        # Apresenta resumo
        print(f"\n{Fore.GREEN}Análise concluída!{Style.RESET_ALL}")
        print(f"\n{Fore.CYAN}Resumo dos Resultados:{Style.RESET_ALL}")
        
        summary = results['summary']
        print(f"• Total de pontos publicados: {summary['total_published_points']}")
        print(f"• Total de valores em falta: {summary['total_missing_values']}")
        print(f"• Percentagem em falta: {summary['missing_percentage']:.2f}%")
        
        # Resumo por folha
        if results['sheet_results']:
            print(f"\n{Fore.CYAN}Resumo por Folha:{Style.RESET_ALL}")
            for sheet_name, sheet_results in results['sheet_results'].items():
                if 'error' not in sheet_results:
                    missing_count = sheet_results.get('missing_count', 0)
                    published_count = sheet_results.get('published_count', 0)
                    missing_percentage = sheet_results.get('missing_percentage', 0.0)
                    
                    status_color = Fore.GREEN if missing_count == 0 else Fore.RED
                    print(f"• {sheet_name}: {status_color}{missing_count} valores em falta{Style.RESET_ALL} "
                          f"({missing_percentage:.2f}% de {published_count} pontos)")
                else:
                    print(f"• {sheet_name}: {Fore.RED}Erro{Style.RESET_ALL} - {sheet_results['error']}")
        
        # Gera relatório
        print(f"\n{Fore.CYAN}Gerando relatório detalhado...{Style.RESET_ALL}")
        report_file = analyzer.generate_missing_values_report(results)
        
        print(f"\n{Fore.GREEN}✅ Relatório gerado com sucesso!{Style.RESET_ALL}")
        print(f"Ficheiro: {report_file}")
        
        # Mostra próximos passos
        if summary['total_missing_values'] > 0:
            print(f"\n{Fore.YELLOW}Próximos Passos:{Style.RESET_ALL}")
            print("1. Revise o relatório Excel gerado para detalhes dos valores em falta")
            print("2. Verifique se os dados estão disponíveis nas fontes originais")
            print("3. Atualize o processo de recriação para incluir os dados em falta")
        else:
            print(f"\n{Fore.GREEN}✅ Excelente!{Style.RESET_ALL} Nenhum valor em falta foi encontrado.")
            print("O dataset recriado contém todos os valores do dataset publicado.")
        
    except Exception as e:
        logger.error(f"Erro durante análise de valores em falta: {e}", exc_info=True)
        print(f"\n{Fore.RED}Erro crítico durante a análise:{Style.RESET_ALL} {str(e)}")
        print("Consulte os logs para detalhes técnicos.") 